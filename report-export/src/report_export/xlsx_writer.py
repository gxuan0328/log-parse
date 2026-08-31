"""Build the 2-sheet, value-only deliverable workbook and resolve its
output filename (design.md §2.3 xlsx_writer, §3.7, §7.1).

Every cell is a literal Python value (`datetime`, `str`, or `int`) --
never a formula string -- reproducing the template's Excel 365
dynamic-array formulas (`UNIQUE`/`FILTER`/`XLOOKUP`/`ANCHORARRAY`/
`COUNTIF`) as pre-computed values (design.md §3.7.1, §8 R5). The
workbook is built entirely in memory (`build_workbook`); `write()`
layers the filesystem concerns on top: same-day filename
disambiguation (design.md §3.7.2), directory creation, and a tmp write +
fsync. The FINAL `os.replace` into the live deliverable name is
deliberately left to the caller (design.md §3.8 S8 vs S9.2: pipeline.py
performs it only after `state.commit()` has already succeeded --
state-first ordering, design.md §4.1).

Every cell (both sheets, header + data) is also center-aligned and
bordered, and every column is auto-fit to whatever data is actually
present this run (design.md §3.7.3-§3.7.5 REQ1) -- a purely cosmetic
presentation layer applied as a post-pass AFTER all values/types/
number_formats/fills are already written, so it can never disturb the
value/type/number_format/fill fidelity asserted in §3.7.6.
"""

from __future__ import annotations

import logging
import os
import unicodedata
from collections.abc import Mapping, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from report_export.errors import WriteError
from report_export.models import ReportRow, StateRecord

__all__ = ["build_workbook", "resolve_filename", "write"]

logger = logging.getLogger(__name__)

_SHEET_RECORDS: Final[str] = "調閱紀錄"
_SHEET_AGGREGATE: Final[str] = "院所分析"

#: design.md §3.1.2: BIRTHDAY is deliberately ABSENT from the delivered
#: workbook. The plaintext date-of-birth is still ingested (csv_reader),
#: transformed (transform), and persisted VERBATIM into the machine-owned
#: records.csv state (state.py / models.StateRecord.birthday) -- only this
#: final Excel projection drops it. PATIENT ID AES therefore closes the
#: 調閱紀錄 sheet at column H (it was column I while BIRTHDAY occupied H).
_RECORDS_HEADER: Final[tuple[str, ...]] = (
    "DATE",
    "TIME",
    "CLIENT IP",
    "SERVER IP",
    "HOSP_ID",
    "HOSP_ABBR",
    "PRSN_ID",
    "PATIENT ID AES",
)
#: design.md §3.1.3/§3.7.4 REQ3: the old single COUNT column split into
#: WEEKLY ACCESS (this run's latest-batch row count) and TOTAL ACCESS
#: (the old COUNT, full-state row count) -- English, space-separated,
#: matching the CLIENT IP/PATIENT ID AES header style.
_AGGREGATE_HEADER: Final[tuple[str, ...]] = (
    "CLIENT IP",
    "HOSP_ID",
    "HOSP_ABBR",
    "WEEKLY ACCESS",
    "TOTAL ACCESS",
)

#: design.md §1.5.6/§3.8 S4: the same two accepted APP_TIME formats as
#: transform.py's contract, duplicated here deliberately -- xlsx_writer
#: is declared to depend only on (openpyxl, models), never transform
#: (design.md §2.3), so the rendering layer re-parses the persisted
#: APP_TIME_ISO string independently rather than importing transform's
#: private constant.
_APP_TIME_FORMATS: Final[tuple[str, ...]] = ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S")

#: design.md §3.7.3: exact number_format strings measured off the real
#: template's 調閱紀錄 sheet (A/B columns).
_DATE_NUMBER_FORMAT: Final[str] = r"yyyy\-mm\-dd;@"
_TIME_NUMBER_FORMAT: Final[str] = r"h:mm:ss;@"
#: design.md §3.1.2/§3.7.3: text columns are hardened to `@` even where the
#: template itself measured `General` -- deliberate, over the template.
_TEXT_NUMBER_FORMAT: Final[str] = "@"

#: design.md §3.7.3, §1.5.9: explicit 8-hex ARGB, FF alpha -- a 6-hex
#: `'FFFF00'` gets stored by openpyxl with alpha=00 (fully transparent).
_HIGHLIGHT_FILL: Final[PatternFill] = PatternFill(fill_type="solid", fgColor="FFFFFF00")
#: design.md §3.7.5: explicit RGB, not a theme index -- reading a theme
#: color's `fgColor.rgb` raises `Values must be of type str` (measured,
#: design.md §1.5.5), and theme colors do not portably travel to a new
#: workbook either way. 8-hex with an explicit `FF` alpha for the same
#: reason as the yellow highlight (design.md §1.5.9): a bare 6-hex
#: `'E2EFDA'` is stored by openpyxl as `'00E2EFDA'` (alpha=00), which
#: would silently defeat the visible-header intent.
_HEADER_FILL: Final[PatternFill] = PatternFill(fill_type="solid", fgColor="FFE2EFDA")
_HEADER_FONT: Final[Font] = Font(bold=True, size=12)

#: design.md §3.7.3-§3.7.5 REQ1a: every cell, header and data, both
#: sheets, centered both horizontally and vertically.
_CENTER: Final[Alignment] = Alignment(horizontal="center", vertical="center")
#: design.md §3.7.3 REQ1b ("所有框線"): thin border on all four sides of
#: every DATA cell, both sheets.
_DATA_BORDER: Final[Border] = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
#: design.md §3.7.5 REQ1c ("粗下框線"): the REQUIRED emphasis is the
#: thick bottom border; thin left/right/top are added so the header
#: row's outline is a continuous grid with the thin-bordered data rows
#: below it, rather than an open top/sides on the table's first row.
_HEADER_BORDER: Final[Border] = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thick"),
)

#: design.md §3.1.3/§3.7.4 REQ3b: rendered in WEEKLY ACCESS whenever an
#: IP has zero rows in the latest batch ("本周無存取紀錄之院所則填入 -").
_WEEKLY_NONE_SENTINEL: Final[str] = "-"

_FILENAME_DATE_FMT: Final[str] = "%Y-%m-%d"
_DELIVERABLE_SUFFIX: Final[str] = "_連線紀錄.xlsx"


# --------------------------------------------------------------------
# build_workbook() -- pure, in-memory (design.md §3.7.1, §3.7.3-§3.7.5)
# --------------------------------------------------------------------


def build_workbook(full_state: Sequence[StateRecord], report_rows: Sequence[ReportRow]) -> Workbook:
    """Build the complete 2-sheet workbook in memory.

    `full_state` order is preserved verbatim onto 調閱紀錄 (existing
    rows first, this run's newly appended batch last, design.md §3.7.3);
    `report_rows` order is preserved verbatim onto 院所分析
    (first-seen order, design.md §3.1.3). Rows whose `batch_id` equals
    the highest BATCH_ID present in `full_state` get the yellow
    highlight (design.md §3.7.3) -- computed fresh from `full_state`
    every call, so a 0-new-records run correctly re-highlights the
    EXISTING latest batch rather than a batch that was never created.

    Exactly 2 sheets, in this order, and the openpyxl default `Sheet`
    is removed (design.md §3.7.1) -- no 紀錄匯入/格式轉換/HOSP_ID對照表,
    and no formula is ever written (every cell gets a literal value).
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_records_sheet(workbook.create_sheet(_SHEET_RECORDS), full_state)
    _write_aggregate_sheet(workbook.create_sheet(_SHEET_AGGREGATE), report_rows)
    return workbook


def _write_records_sheet(sheet: Worksheet, full_state: Sequence[StateRecord]) -> None:
    _write_header(sheet, _RECORDS_HEADER)
    max_batch_id = max((record.batch_id for record in full_state), default=0)
    for row_idx, record in enumerate(full_state, start=2):
        _write_record_row(sheet, row_idx, record, highlight=record.batch_id == max_batch_id)
    _autofit(sheet, _RECORDS_HEADER)


def _write_record_row(
    sheet: Worksheet, row_idx: int, record: StateRecord, *, highlight: bool
) -> None:
    # design.md §3.1.2/§3.7.3: A(DATE) and B(TIME) hold the SAME datetime
    # object -- only the number_format differs -- so the round-trip
    # value (including microseconds) is identical on both cells.
    app_time = _parse_app_time_iso(record.app_time_iso)

    date_cell = sheet.cell(row=row_idx, column=1, value=app_time)
    date_cell.number_format = _DATE_NUMBER_FORMAT
    time_cell = sheet.cell(row=row_idx, column=2, value=app_time)
    time_cell.number_format = _TIME_NUMBER_FORMAT
    row_cells = [date_cell, time_cell]

    # record.birthday is intentionally NOT projected here (design.md
    # §3.1.2): it remains in state/records.csv but never reaches the
    # deliverable, so PATIENT ID AES shifts left one column (I -> H).
    text_values = (
        record.client_ip,
        record.server_ip,
        record.hosp_id,
        record.hosp_abbr,
        record.prsn_id,
        record.patient_id_aes,
    )
    for column, value in enumerate(text_values, start=3):
        cell = sheet.cell(row=row_idx, column=column, value=value)
        cell.number_format = _TEXT_NUMBER_FORMAT
        row_cells.append(cell)

    # design.md §3.7.3 REQ1a/b: center + thin 4-side border on every DATA
    # cell, unconditionally; the yellow highlight (latest batch only)
    # layers on top of -- never instead of -- that grid/alignment.
    for cell in row_cells:
        cell.alignment = _CENTER
        cell.border = _DATA_BORDER
        if highlight:
            cell.fill = _HIGHLIGHT_FILL


def _write_aggregate_sheet(sheet: Worksheet, report_rows: Sequence[ReportRow]) -> None:
    _write_header(sheet, _AGGREGATE_HEADER)
    for row_idx, report_row in enumerate(report_rows, start=2):
        text_values = (report_row.client_ip, report_row.hosp_id, report_row.hosp_abbr)
        row_cells = [
            sheet.cell(row=row_idx, column=column, value=value)
            for column, value in enumerate(text_values, start=1)
        ]
        for cell in row_cells:
            cell.number_format = _TEXT_NUMBER_FORMAT

        # design.md §3.1.3/§3.7.4 REQ3b-c: WEEKLY ACCESS renders the "-"
        # sentinel (str, hardened `@`) when this IP had no rows in the
        # latest batch; otherwise the int is left at the default
        # `General` number_format, same as TOTAL ACCESS always is.
        if report_row.weekly_access == 0:
            weekly_cell = sheet.cell(row=row_idx, column=4, value=_WEEKLY_NONE_SENTINEL)
            weekly_cell.number_format = _TEXT_NUMBER_FORMAT
        else:
            weekly_cell = sheet.cell(row=row_idx, column=4, value=report_row.weekly_access)
        total_cell = sheet.cell(row=row_idx, column=5, value=report_row.total_access)
        row_cells += [weekly_cell, total_cell]

        for cell in row_cells:
            cell.alignment = _CENTER
            cell.border = _DATA_BORDER
    _autofit(sheet, _AGGREGATE_HEADER)


def _write_header(sheet: Worksheet, header: tuple[str, ...]) -> None:
    for column, title in enumerate(header, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _HEADER_BORDER


def _parse_app_time_iso(raw: str) -> datetime:
    for fmt in _APP_TIME_FORMATS:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    # Unreachable through the pipeline's own call graph: every
    # APP_TIME_ISO persisted into state already passed
    # transform._validate_app_time's identical format check before
    # being written (design.md §3.8 S5-S7) -- kept as a fail-loud guard
    # against a hand-edited or otherwise-corrupted records.csv
    # (design.md §4.7.7 R7: records.csv is machine-owned, never hand-edited,
    # but this module must not silently mis-render it if that happens).
    raise WriteError(f"cannot render APP_TIME_ISO as a datetime: {raw!r}")


# --------------------------------------------------------------------
# Column auto-fit -- dynamic, from THIS run's actual data (design.md
# §3.7.3/§3.7.4 REQ1d). Runs as a post-pass after every cell already holds
# its final value/number_format, so it only ever READS what earlier
# code wrote -- it can neither disturb nor depend on write order.
# --------------------------------------------------------------------


def _display_width(text: str) -> int:
    """A tiny wcwidth-style helper (design.md §3.7.3 REQ1d): CJK/full-width
    code points (`unicodedata.east_asian_width` in `{"W", "F"}`) count
    as 2 display columns, everything else as 1 -- so a 4-character
    hospital name like `門諾醫院` sizes as 8, not 4, and a column holding
    it is not visually truncated even though `len()` would say 4.
    """
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1 for ch in text)


def _rendered(value: object, number_format: str) -> str:
    """The literal characters a cell DISPLAYS (design.md §3.7.3 REQ1d) --
    distinct from `cell.value`'s Python type:

        datetime + a DATE-style format (`"yyyy"` in `number_format`)
            -> `"YYYY-MM-DD"` (always 10 characters)
        datetime + anything else (the TIME column)
            -> `"HH:MM:SS"` (always 8 characters, no sub-second digits)
        None (an unmapped HOSP_ABBR written as `""`, which openpyxl
        normalizes to `None` on read-back, design.md §1.5.9)
            -> `""`
        anything else (str, int, including the WEEKLY "-" sentinel)
            -> `str(value)` (an int renders as its digit string)
    """
    if isinstance(value, datetime):
        if "yyyy" in number_format:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%H:%M:%S")
    if value is None:
        # Defensive only: every cell _autofit() ever reads back was
        # JUST written by _write_record_row/_write_aggregate_sheet in
        # this same build_workbook() call, and neither ever assigns a
        # bare `None` (an unmapped HOSP_ABBR is written as `""`, which
        # only normalizes to `None` on a save+reload round-trip that
        # has not happened yet at autofit time, design.md §1.5.9) -- so
        # this branch is unreachable through the current call graph.
        # Kept anyway so `_rendered()` stays a total, honest function
        # of "what would this cell display" for any caller/future cell
        # kind, not just the ones this module happens to write today.
        return ""  # pragma: no cover
    return str(value)


def _autofit(sheet: Worksheet, header: Sequence[str]) -> None:
    """Set every column's width to fit whatever data THIS sheet
    actually holds this run (design.md §3.7.3/§3.7.4 REQ1d) -- dynamic,
    never a hardcoded constant. Per column: `width = max(display_width
    of the header text, display_width of every data cell's RENDERED
    value) * 1.2` -- the header participates in the max so a wide
    header (e.g. "WEEKLY ACCESS") is never truncated by narrower data
    (design.md §3.7.5: 確保清晰的資訊檢視). A header-only (0 data row) sheet
    still gets a sensible width, since the header-only term is always
    present in the max.

    `round(..., 2)` (design.md §3.7.3): 2 decimal places is exactly
    enough precision for `display_width * 1.2` to always be exact (the
    only fractional parts that formula can ever produce are `.0`, `.2`,
    `.4`, `.6`, `.8`), so this round-trips through openpyxl without
    float drift and tests can assert exact width floats.
    """
    for col_idx, title in enumerate(header, start=1):
        widths = [_display_width(title)]
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            widths.append(_display_width(_rendered(cell.value, cell.number_format)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = round(max(widths) * 1.2, 2)


# --------------------------------------------------------------------
# Filename resolution -- same-day disambiguation (design.md §3.7.2)
# --------------------------------------------------------------------


def resolve_filename(
    *,
    run_date: date,
    out_dir: Path,
    today_runs: Sequence[Mapping[str, object]],
    input_sha256: str,
) -> str:
    """Resolve the deliverable's filename for this run (design.md §3.7.2).

    `today_runs` is every `runs.jsonl` record already logged for
    `run_date` (chronological order), fetched by the caller
    (pipeline.py, via `state.read_runs_for_date`) -- this function only
    reasons over that already-filtered slice, never touches the
    filesystem for `runs.jsonl` itself, keeping xlsx_writer's
    dependency to (openpyxl, models) as declared (design.md §2.3).

    The base name `{run_date}_連線紀錄.xlsx` is used whenever it does
    not already exist in `out_dir` -- the common case (first ingest of
    the day). If it DOES exist: an identical `input_sha256` to the last
    logged run today means an idempotent re-ingest -- reuse the same
    name (deterministic overwrite, design.md §4.1); a DIFFERENT
    `input_sha256` means a genuinely distinct second batch the same day
    -- append `_{seq:02d}` (design.md §3.7.2, §6-18).
    """
    base_name = f"{run_date.strftime(_FILENAME_DATE_FMT)}{_DELIVERABLE_SUFFIX}"
    if not (out_dir / base_name).exists():
        return base_name
    if today_runs and today_runs[-1].get("input_sha256") == input_sha256:
        return base_name
    # The base file already existing proves at least one earlier batch
    # landed today even if `today_runs` has no record of it (e.g. a
    # lost/reset runs.jsonl) -- never silently reuse/clobber it without
    # a confirmed sha256 match; the disambiguated suffix always starts
    # at 2, never 1 (a lone "_01" would be a nonsensical sibling of the
    # unsuffixed base name that is never itself produced).
    seq = max(len(today_runs), 1) + 1
    stem = base_name.removesuffix(".xlsx")
    return f"{stem}_{seq:02d}.xlsx"


# --------------------------------------------------------------------
# write() -- build + persist to `.tmp`, fsync (design.md §3.8 S8)
# --------------------------------------------------------------------


def write(
    out_dir: Path,
    full_state: Sequence[StateRecord],
    report_rows: Sequence[ReportRow],
    *,
    run_date: date,
    today_runs: Sequence[Mapping[str, object]],
    input_sha256: str,
) -> tuple[Path, Path]:
    """Build the workbook and persist it to `{filename}.tmp` under
    `out_dir`, fsync'd (design.md §3.8 S8).

    Returns `(tmp_path, final_path)` WITHOUT renaming `tmp_path` into
    place -- the caller (pipeline.py S9.2) performs that `os.replace`
    only after `state.commit()` has already succeeded, so the
    deliverable never becomes visible ahead of the state it is a
    projection of (design.md §4.1 state-first ordering).

    Raises:
        WriteError (exit 5): `out_dir` cannot be prepared, or the
            workbook cannot be built/saved/fsync'd.
    """
    _ensure_out_dir(out_dir)
    filename = resolve_filename(
        run_date=run_date, out_dir=out_dir, today_runs=today_runs, input_sha256=input_sha256
    )
    final_path = out_dir / filename
    tmp_path = out_dir / f"{filename}.tmp"

    workbook = build_workbook(full_state, report_rows)
    try:
        workbook.save(str(tmp_path))
        os.chmod(tmp_path, 0o600)
        _fsync_path(tmp_path)
    except OSError as exc:
        raise WriteError(f"cannot write deliverable workbook: {tmp_path} ({exc})") from exc
    return tmp_path, final_path


def _ensure_out_dir(out_dir: Path) -> None:
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        os.chmod(out_dir, 0o700)
    except OSError as exc:
        raise WriteError(f"cannot prepare out_dir: {out_dir} ({exc})") from exc


def _fsync_path(path: Path) -> None:
    fd = os.open(path, os.O_RDONLY)
    try:
        os.fsync(fd)
    finally:
        os.close(fd)
