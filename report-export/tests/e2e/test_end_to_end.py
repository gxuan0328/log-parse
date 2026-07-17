"""End-to-end acceptance tests (design.md §7.2, §7.3).

Drives the full S0-S9 pipeline (`pipeline.run()`) through real
filesystem I/O -- `tmp_path` state/output directories, real `Config`
objects, real `openpyxl` round-trips -- rather than unit-testing
individual stages in isolation (that is `tests/unit/`'s job). Two
inputs anchor almost every scenario here:

  * `template/source-log.csv`      the 22-row, CRLF, real anchor
                                    dataset (design.md §1.5.1/§1.5.2/§1.5.6),
                                    referenced directly rather than
                                    duplicated into tests/fixtures/ --
                                    design.md §7.3 explicitly permits
                                    "直接引用或複製", and duplicating a
                                    22-row byte-exact fixture would
                                    create a second source of truth
                                    that could silently drift from the
                                    committed baseline (project
                                    CLAUDE.md §2 "Single source of
                                    truth").
  * `reference/hosp_id_map.csv.gz` the real, bundled 93,781-row lookup
                                    table (design.md §3.3) -- the same
                                    file the production Docker image
                                    ships.

Every `pipeline.run()` call here injects a FIXED `run_date` (design.md
§3.9.1 test seam) -- never `date.today()` -- so filenames, batch
sequencing, and same-day disambiguation are all deterministic.

Two golden-master baselines, generated once by actually running this
same pipeline against this same anchor input and checked into
`tests/fixtures/`, pin the exact E2E-1 output as a byte-for-byte /
cell-level regression baseline:

  * `expected_records_e2e1.csv`      the verbatim `records.csv` body
                                      E2E-1 produces. `records.csv`
                                      embeds no wall-clock timestamps
                                      anywhere (design.md §3.5.2), so this
                                      is safe to compare as raw bytes,
                                      not just "a value set".
  * `expected_deliverable_e2e1.json` a JSON snapshot of every cell in
                                      both deliverable sheets --
                                      `(value, data_type, number_format,
                                      fill)` -- NEVER compared as xlsx
                                      bytes (the zip container embeds a
                                      non-deterministic docProps
                                      timestamp on every save, design.md
                                      §7.2 E2E-2).
"""

from __future__ import annotations

import csv
import gzip
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from report_export import csv_reader, lookup, pipeline, state, transform
from report_export.config import Config
from report_export.errors import StateIntegrityError
from report_export.models import TransformedRecord

# --------------------------------------------------------------------
# Shared paths + constants
# --------------------------------------------------------------------

_REPO_ROOT: Path = Path(__file__).resolve().parents[2]
_FIXTURES_DIR: Path = Path(__file__).resolve().parent.parent / "fixtures"

#: design.md §7.3: the main e2e input is REFERENCED directly from the
#: checked-in Phase-0 baseline, never duplicated.
_SOURCE_LOG: Path = _REPO_ROOT / "template" / "source-log.csv"
_REAL_REFERENCE: Path = _REPO_ROOT / "reference" / "hosp_id_map.csv.gz"

_BATCH_NEW: Path = _FIXTURES_DIR / "batch_new.csv"
_EMPTY: Path = _FIXTURES_DIR / "empty.csv"
_ALL_NONNORMAL: Path = _FIXTURES_DIR / "all_nonnormal.csv"
_STATUS_MIXED_CASE: Path = _FIXTURES_DIR / "status_mixed_case.csv"
_HOSP_MAP_SMALL: Path = _FIXTURES_DIR / "hosp_map_small.csv"

_EXPECTED_RECORDS_E2E1: Path = _FIXTURES_DIR / "expected_records_e2e1.csv"
_EXPECTED_DELIVERABLE_E2E1: Path = _FIXTURES_DIR / "expected_deliverable_e2e1.json"

#: design.md §7.3: fixed injected run_date, never date.today(), for
#: determinism. Three distinct dates stand in for three distinct weekly
#: runs so same-day filename disambiguation (design.md §3.7.2) never
#: incidentally triggers where a test does not intend to exercise it.
_RUN_DATE_1: date = date(2026, 7, 15)
_RUN_DATE_2: date = date(2026, 7, 22)
_RUN_DATE_3: date = date(2026, 7, 29)

#: design.md §1.5.2 anchor: 院所分析 WEEKLY/TOTAL ACCESS columns, first-seen
#: order -- E2E-1 is a single-batch state, so WEEKLY == TOTAL == these
#: values for every IP (design.md §3.1.3 REQ3), same as the old COUNT.
_ANCHOR_COUNTS: tuple[int, ...] = (1, 1, 1, 1, 1, 1, 1, 7, 1, 1)

#: design.md §3.7.3/§3.7.4 REQ1d anchors: round(display_width * 1.2, 2)
#: for the E2E-1 anchor state, both sheets (see the docstring on
#: `_assert_column_widths` for the formula).
_RECORDS_SHEET_WIDTHS: dict[str, float] = {
    "A": 12.0,
    "B": 9.6,
    "C": 16.8,
    "D": 12.0,
    "E": 12.0,
    "F": 12.0,
    "G": 38.4,
    "H": 9.6,
    "I": 38.4,
}
_AGGREGATE_SHEET_WIDTHS: dict[str, float] = {
    "A": 16.8,
    "B": 12.0,
    "C": 12.0,
    "D": 15.6,
    "E": 14.4,
}


def _config(tmp_path: Path, input_path: Path) -> Config:
    return Config(input_path=input_path, state_dir=tmp_path / "state", out_dir=tmp_path / "output")


def _gzip_reference(csv_path: Path, dest_dir: Path) -> Path:
    """gzip-compress a human-readable fixture CSV into `dest_dir` so it
    can be passed as `pipeline.run(reference_path=...)` --
    `lookup.load()` requires gzip (design.md §3.3) -- while the
    checked-in fixture itself stays plain-text and git-diffable
    (design.md §3.5.1 CSV ethos).
    """
    dest = dest_dir / f"{csv_path.stem}.csv.gz"
    with csv_path.open("rb") as src, gzip.open(dest, "wb") as dst:
        dst.write(src.read())
    return dest


# --------------------------------------------------------------------
# Workbook cell-level snapshot: (value, data_type, number_format, fill)
# -- design.md §7.2 E2E-2: "儲存格值 + data_type + number_format + fill
# 相等(非 bytes)". datetimes render as isoformat strings so the whole
# structure is plain-JSON-comparable, and diffable against the
# checked-in golden-master baseline.
# --------------------------------------------------------------------


def _cell_snapshot(cell: Any) -> dict[str, Any]:
    value = cell.value
    if isinstance(value, datetime):
        value = value.isoformat()
    fill = cell.fill
    fill_info: dict[str, Any] | None = None
    if fill is not None and fill.patternType == "solid":
        fill_info = {"patternType": fill.patternType, "fgColor": fill.fgColor.rgb}
    return {
        "value": value,
        "data_type": cell.data_type,
        "number_format": cell.number_format,
        "fill": fill_info,
    }


def _sheet_snapshot(sheet: Worksheet) -> list[list[dict[str, Any]]]:
    return [[_cell_snapshot(cell) for cell in row] for row in sheet.iter_rows()]


def _workbook_snapshot(path: Path) -> dict[str, list[list[dict[str, Any]]]]:
    workbook = load_workbook(path)
    return {name: _sheet_snapshot(workbook[name]) for name in workbook.sheetnames}


def _load_expected_deliverable_snapshot() -> dict[str, list[list[dict[str, Any]]]]:
    with _EXPECTED_DELIVERABLE_E2E1.open(encoding="utf-8") as fh:
        snapshot: dict[str, list[list[dict[str, Any]]]] = json.load(fh)
        return snapshot


def _highlighted_rows(sheet: Worksheet) -> list[int]:
    """1-indexed data-row numbers (excludes the header) whose column-A
    cell carries the solid yellow highlight fill (design.md §3.7.3).
    """
    return [
        row_idx
        for row_idx in range(2, sheet.max_row + 1)
        if sheet.cell(row=row_idx, column=1).fill.patternType == "solid"
    ]


def _assert_no_formulas(workbook: Any) -> None:
    """design.md §3.7.1/§3.7.6: every cell is a literal value -- never a
    formula string (openpyxl's `data_type == 'f'` marks a formula).
    """
    for sheet_name in workbook.sheetnames:
        for row in workbook[sheet_name].iter_rows():
            for cell in row:
                location = f"{sheet_name}!{cell.coordinate}"
                assert cell.data_type != "f", f"unexpected formula at {location}"


def _assert_uniform_grid_and_alignment(workbook: Any) -> None:
    """design.md §3.7.3-§3.7.6 REQ1: every cell (both sheets) is centered;
    every DATA cell has a thin 4-side border; every HEADER cell has a
    thick bottom border (the REQUIRED emphasis) plus thin left/right/
    top (cosmetic grid continuity with the data rows below it).
    """
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for cell in sheet[1]:
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
            assert cell.border.bottom.style == "thick"
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.top.style == "thin"
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                assert cell.alignment.horizontal == "center"
                assert cell.alignment.vertical == "center"
                assert cell.border.left.style == "thin"
                assert cell.border.right.style == "thin"
                assert cell.border.top.style == "thin"
                assert cell.border.bottom.style == "thin"


def _assert_column_widths(sheet: Worksheet, expected: dict[str, float]) -> None:
    """design.md §3.7.3/§3.7.4 REQ1d: `width == round(max(display_width(
    header), max(display_width(rendered data))) * 1.2, 2)` per column
    -- asserted here as an exact expected-value dict per sheet, computed
    once off the real E2E-1 anchor data (design.md §1.5.1/§1.5.2).
    """
    for column, width in expected.items():
        assert sheet.column_dimensions[column].width == width


def _tail_line(records_path: Path) -> str:
    return records_path.read_text(encoding="utf-8").splitlines()[-1]


def _corrupt_tail_sha(path: Path) -> None:
    """Flip the `#META` tail's `sha256=` value so it no longer matches
    the file's own body -- design.md §3.5.4 case 4 (tail present but
    mismatched).
    """
    lines = path.read_text(encoding="utf-8").splitlines()
    assert lines[-1].startswith("#META")
    lines[-1] = lines[-1].replace("sha256=", "sha256=deadbeef")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# --------------------------------------------------------------------
# records.csv body -> expected TransformedRecord list. design.md §3.5.2's
# 10 state columns minus the internal BATCH_ID key are EXACTLY
# TransformedRecord's 9 fields, in the same order -- so the same
# checked-in golden master serves both the state-level (E2E-1) and the
# pure-transform-level (invariant) assertions below, one source of
# truth for "what are the 19 expected rows" (project CLAUDE.md §2).
# --------------------------------------------------------------------


def _load_expected_transformed_records() -> list[TransformedRecord]:
    text = _EXPECTED_RECORDS_E2E1.read_text(encoding="utf-8")
    body_lines = [line for line in text.splitlines() if not line.startswith("#META")]
    reader = csv.reader(body_lines)
    next(reader)  # header: BATCH_ID,REQUEST_ID,APP_TIME_ISO,CLIENT_IP,...
    return [
        TransformedRecord(
            request_id=row[1],
            app_time_iso=row[2],
            client_ip=row[3],
            server_ip=row[4],
            hosp_id=row[5],
            hosp_abbr=row[6],
            prsn_id=row[7],
            birthday=row[8],
            patient_id_aes=row[9],
        )
        for row in reader
    ]


# --------------------------------------------------------------------
# Fixture integrity -- the demo input carries no log-parse test-host IP
# --------------------------------------------------------------------

#: design.md §1.5.6: the four log-parse test-host client IPs are
#: pre-filtered upstream (analyze_access --test-hosts exclude) and were
#: struck from template/source-log.csv in the test-host re-baseline.
#: None may reappear -- a test-host IP here would misattribute internal
#: QA/probe traffic to a real 院所 in the 院所分析 aggregate.
_TEST_HOST_IPS: frozenset[str] = frozenset(
    {"192.168.117.90", "192.168.105.149", "192.168.117.73", "192.168.117.104"}
)


def test_source_log_demo_input_excludes_test_host_ips() -> None:
    with _SOURCE_LOG.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    header, data = rows[0], rows[1:]
    ip_column = header.index("CLIENT_IP")
    present = {row[ip_column] for row in data if row}
    leaked = present & _TEST_HOST_IPS
    assert not leaked, f"log-parse test-host IP leaked into demo fixture: {sorted(leaked)}"


# --------------------------------------------------------------------
# E2E-1 -- empty state + source-log.csv(22) (design.md §7.2, §1.5.1, §1.5.2)
# --------------------------------------------------------------------


def test_e2e1_empty_state_first_ingest_matches_anchors(tmp_path: Path) -> None:
    config = _config(tmp_path, _SOURCE_LOG)

    summary = pipeline.run(config, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)

    assert summary.rows_in == 22
    assert summary.normal == 16
    assert summary.dropped_nonnormal == 6
    assert summary.new_records == 16
    assert summary.state_total == 16
    assert summary.unique_ips == 10
    assert summary.batch_seq == 1
    assert summary.unmapped_hosp_ids == 0

    # State: exactly the 16 NORMAL rows, all BATCH_ID=1, byte-for-byte
    # identical to the checked-in golden master (records.csv embeds no
    # timestamps -- fully reproducible, design.md §3.5.2).
    records_path = config.state_dir / "records.csv"
    assert records_path.read_bytes() == _EXPECTED_RECORDS_E2E1.read_bytes()
    result = state.load(config.state_dir)
    assert len(result.existing) == 16
    assert {r.batch_id for r in result.existing} == {1}

    workbook = load_workbook(summary.deliverable)
    assert workbook.sheetnames == ["調閱紀錄", "院所分析"]  # exactly 2 sheets
    _assert_no_formulas(workbook)  # zero formulas

    records_sheet = workbook["調閱紀錄"]
    assert records_sheet.max_row == 17  # header + 16 NORMAL rows
    # ALL 16 data rows are the sole (and thus latest) batch -> all yellow.
    assert _highlighted_rows(records_sheet) == list(range(2, 18))
    assert records_sheet["A2"].fill.fgColor.rgb == "FFFFFF00"  # explicit 8-hex ARGB

    # A2==B2: the SAME datetime, only number_format differs.
    assert isinstance(records_sheet["A2"].value, datetime)
    assert records_sheet["A2"].value == records_sheet["B2"].value
    assert "yyyy" in records_sheet["A2"].number_format
    assert "h:mm" in records_sheet["B2"].number_format

    # Leading-zero HOSP_ID (0937010019) survives as `str`, hardened `@`.
    hosp_id_row = next(
        row_idx
        for row_idx in range(2, 18)
        if records_sheet.cell(row=row_idx, column=5).value == "0937010019"
    )
    hosp_id_cell = records_sheet.cell(row=hosp_id_row, column=5)
    assert hosp_id_cell.value == "0937010019"
    assert isinstance(hosp_id_cell.value, str)
    assert hosp_id_cell.data_type == "s"
    assert hosp_id_cell.number_format == "@"

    agg_sheet = workbook["院所分析"]
    assert agg_sheet.max_row == 11  # header + 10 unique IPs
    # design.md §3.1.3 REQ3: col D (idx 3) = WEEKLY ACCESS, col E (idx 4)
    # = TOTAL ACCESS. E2E-1 is a single-batch state (all BATCH_ID=1),
    # so WEEKLY == TOTAL == the old COUNT for every IP -- no "-" cell.
    weekly = [row[3].value for row in agg_sheet.iter_rows(min_row=2)]
    total = [row[4].value for row in agg_sheet.iter_rows(min_row=2)]
    assert weekly == list(_ANCHOR_COUNTS)
    assert total == list(_ANCHOR_COUNTS)
    assert sum(weekly) == 16
    assert sum(total) == 16
    assert all(isinstance(v, int) for v in weekly)  # never "-" in a single-batch run

    shuchuan = next(row for row in agg_sheet.iter_rows(min_row=2) if row[1].value == "0937010019")
    assert shuchuan[2].value == "秀傳醫院"
    assert shuchuan[3].value == 7
    assert shuchuan[4].value == 7
    orphan_ip_row = next(
        row for row in agg_sheet.iter_rows(min_row=2) if row[0].value == "10.243.129.44"
    )
    assert orphan_ip_row[3].value == 1  # proves the ORPHAN row (same IP) was excluded
    assert orphan_ip_row[4].value == 1

    # REQ1: center alignment + thin/thick borders (uniform, both
    # sheets) + auto-fit column widths off the real anchor data.
    _assert_uniform_grid_and_alignment(workbook)
    _assert_column_widths(records_sheet, _RECORDS_SHEET_WIDTHS)
    _assert_column_widths(agg_sheet, _AGGREGATE_SHEET_WIDTHS)

    # Full cell-level (value + data_type + number_format + fill) golden master.
    assert _workbook_snapshot(Path(summary.deliverable)) == _load_expected_deliverable_snapshot()


# --------------------------------------------------------------------
# E2E-2 -- idempotent rerun (design.md §7.2, §4.1)
# --------------------------------------------------------------------


def test_e2e2_idempotent_rerun_is_a_true_no_op(tmp_path: Path) -> None:
    config = _config(tmp_path, _SOURCE_LOG)
    first = pipeline.run(config, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)
    records_path = config.state_dir / "records.csv"
    state_bytes_before = records_path.read_bytes()
    first_snapshot = _workbook_snapshot(Path(first.deliverable))

    second = pipeline.run(config, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)

    assert second.new_records == 0
    assert second.skipped_cross_state == 16
    assert second.state_total == 16
    assert second.batch_seq == first.batch_seq == 1
    assert second.deliverable == first.deliverable  # same-day, same sha256 -> same filename

    # State tail/bytes literally unchanged -- commit() never ran a second time.
    assert records_path.read_bytes() == state_bytes_before
    assert not (config.state_dir / "records.csv.bak").exists()

    # Two deliverables equal by (value + data_type + number_format + fill),
    # NEVER compared as raw bytes (xlsx zip/docProps timestamps differ).
    second_snapshot = _workbook_snapshot(Path(second.deliverable))
    assert second_snapshot == first_snapshot
    assert second_snapshot == _load_expected_deliverable_snapshot()

    records_sheet = load_workbook(second.deliverable)["調閱紀錄"]
    assert _highlighted_rows(records_sheet) == list(range(2, 18))  # all 16, still batch 1


# --------------------------------------------------------------------
# E2E-3 -- new batch, per-run yellow reset (design.md §7.2, §4.1, §3.8 S7)
# --------------------------------------------------------------------


def test_e2e3_new_batch_appends_batch_2_and_recomputes_aggregate(tmp_path: Path) -> None:
    config1 = _config(tmp_path, _SOURCE_LOG)
    pipeline.run(config1, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)

    config2 = Config(input_path=_BATCH_NEW, state_dir=config1.state_dir, out_dir=config1.out_dir)
    summary = pipeline.run(config2, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE)

    n_new = 2  # batch_new.csv's row count -- all brand-new REQUEST_IDs
    assert summary.new_records == n_new
    assert summary.state_total == 16 + n_new
    assert summary.batch_seq == 2

    result = state.load(config1.state_dir)
    assert sum(1 for r in result.existing if r.batch_id == 1) == 16
    assert sum(1 for r in result.existing if r.batch_id == 2) == n_new

    workbook = load_workbook(summary.deliverable)
    records_sheet = workbook["調閱紀錄"]
    assert records_sheet.max_row == 1 + summary.state_total
    assert _highlighted_rows(records_sheet) == [18, 19]  # exactly the N new batch-2 rows
    for row_idx in range(2, 18):  # all 16 batch-1 rows: no fill
        assert records_sheet.cell(row=row_idx, column=1).fill.patternType is None

    agg_sheet = workbook["院所分析"]
    assert agg_sheet.max_row == 12  # 10 existing + 1 brand-new IP
    # design.md §3.1.3 REQ3: col D (idx 3) = WEEKLY ACCESS (this batch's
    # rows only), col E (idx 4) = TOTAL ACCESS (unchanged old COUNT
    # semantics). batch_new.csv's 2 rows each hit an IP exactly once.
    rows_by_ip = {row[0].value: row for row in agg_sheet.iter_rows(min_row=2)}
    assert rows_by_ip["10.245.1.125"][3].value == 1  # WEEKLY: this batch's 1 new row
    assert rows_by_ip["10.245.1.125"][4].value == 8  # TOTAL: 秀傳醫院 7 -> 8
    assert rows_by_ip["10.250.77.10"][3].value == 1  # brand-new 11th IP: weekly == total
    assert rows_by_ip["10.250.77.10"][4].value == 1
    assert rows_by_ip["10.250.77.10"][1].value == "1301170017"
    assert rows_by_ip["10.250.77.10"][2].value == "台北醫大"
    # An older-only (seed-batch-1-only) IP has no rows in the new
    # latest batch -> WEEKLY ACCESS renders "-"; TOTAL ACCESS unchanged.
    assert rows_by_ip["10.243.129.44"][3].value == "-"
    assert rows_by_ip["10.243.129.44"][4].value == 1


# --------------------------------------------------------------------
# E2E-4 -- overlapping re-import after a second batch (design.md §7.2)
# --------------------------------------------------------------------


def test_e2e4_overlap_reimport_after_second_batch_keeps_batch_2_highlighted(
    tmp_path: Path,
) -> None:
    config1 = _config(tmp_path, _SOURCE_LOG)
    pipeline.run(config1, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)
    config2 = Config(input_path=_BATCH_NEW, state_dir=config1.state_dir, out_dir=config1.out_dir)
    pipeline.run(config2, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE)

    config3 = Config(input_path=_SOURCE_LOG, state_dir=config1.state_dir, out_dir=config1.out_dir)
    summary = pipeline.run(config3, run_date=_RUN_DATE_3, reference_path=_REAL_REFERENCE)

    assert summary.new_records == 0
    assert summary.skipped_cross_state == 16  # the reimported 16 all dedup away
    assert summary.state_total == 18  # unchanged: 16 (batch1) + 2 (batch2)
    assert summary.batch_seq == 2  # unchanged: the latest REAL batch is still batch 2

    result = state.load(config1.state_dir)
    assert max(r.batch_id for r in result.existing) == 2

    records_sheet = load_workbook(summary.deliverable)["調閱紀錄"]
    assert _highlighted_rows(records_sheet) == [18, 19]  # still batch 2's rows only


# --------------------------------------------------------------------
# E2E-5 -- deliverable rebuild after loss (design.md §7.2, §4.1)
# --------------------------------------------------------------------


def test_e2e5_deliverable_rebuild_after_loss_reruns_latest_input(tmp_path: Path) -> None:
    config1 = _config(tmp_path, _SOURCE_LOG)
    pipeline.run(config1, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)
    config2 = Config(input_path=_BATCH_NEW, state_dir=config1.state_dir, out_dir=config1.out_dir)
    committed = pipeline.run(config2, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE)

    deliverable_path = Path(committed.deliverable)
    assert deliverable_path.exists()
    deliverable_path.unlink()
    assert not deliverable_path.exists()

    # No special flag/command exists for this (design.md §4.1) -- simply
    # rerunning the most recent input regenerates the deliverable.
    rebuilt = pipeline.run(config2, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE)

    assert rebuilt.new_records == 0  # already-committed batch -- pure regeneration
    assert rebuilt.state_total == committed.state_total
    assert rebuilt.deliverable == committed.deliverable
    assert Path(rebuilt.deliverable).exists()

    records_sheet = load_workbook(rebuilt.deliverable)["調閱紀錄"]
    assert _highlighted_rows(records_sheet) == [18, 19]  # correct latest-batch highlight


# --------------------------------------------------------------------
# E2E-6 -- crash-tolerant state (design.md §7.2, §3.5.4)
# --------------------------------------------------------------------


def test_e2e6_missing_tail_is_non_fatal_and_gets_backfilled(tmp_path: Path) -> None:
    config = _config(tmp_path, _SOURCE_LOG)
    pipeline.run(config, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)
    records_path = config.state_dir / "records.csv"
    assert _tail_line(records_path).startswith("#META")

    # Case 3 (design.md §3.5.4): strip the #META tail entirely, simulating
    # a hand-edited / pre-tail-schema records.csv.
    body_only_lines = records_path.read_text(encoding="utf-8").splitlines()[:-1]
    records_path.write_text("\n".join(body_only_lines) + "\n", encoding="utf-8")

    result = state.load(config.state_dir)  # must NOT brick
    assert len(result.existing) == 16
    assert result.max_batch_seq == 1

    # A 0-new rerun of the SAME input deliberately skips state.commit()
    # entirely (design.md §3.8 S9: idempotent, never touches .bak) -- so it
    # does NOT itself backfill the tail. Ingesting a genuinely new batch
    # does trigger a real commit(), which unconditionally (re)writes a
    # correct tail as part of its normal write path.
    rerun = pipeline.run(config, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)
    assert rerun.new_records == 0  # the 19 REQUEST_IDs were still visible to dedup
    assert not _tail_line(records_path).startswith("#META")  # commit() never ran -- still missing

    new_batch_config = Config(
        input_path=_BATCH_NEW, state_dir=config.state_dir, out_dir=config.out_dir
    )
    committing_rerun = pipeline.run(
        new_batch_config, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE
    )
    assert committing_rerun.new_records == 2
    assert _tail_line(records_path).startswith("#META")  # backfilled by this real commit


def test_e2e6_corrupt_tail_recovers_from_bak(tmp_path: Path) -> None:
    config1 = _config(tmp_path, _SOURCE_LOG)
    pipeline.run(config1, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)
    config2 = Config(input_path=_BATCH_NEW, state_dir=config1.state_dir, out_dir=config1.out_dir)
    pipeline.run(config2, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE)  # creates .bak

    records_path = config1.state_dir / "records.csv"
    bak_path = config1.state_dir / "records.csv.bak"
    assert bak_path.is_file()

    # Case 4 (design.md §3.5.4): corrupt the CURRENT tail's sha256 --
    # records.csv no longer self-validates, but a valid .bak exists.
    _corrupt_tail_sha(records_path)

    result = state.load(config1.state_dir)  # must NOT brick -- recovers from .bak
    assert len(result.existing) == 16  # the older, still-valid .bak snapshot (pre-batch-2)
    assert result.max_batch_seq == 1

    # The full pipeline also tolerates it end-to-end. batch_new.csv's 2
    # REQUEST_IDs are no longer visible in the rolled-back (.bak) state,
    # so this rerun legitimately re-ingests them as batch 2 again --
    # self-healing, never a brick.
    rerun = pipeline.run(config2, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE)
    assert rerun.new_records == 2
    assert rerun.state_total == 18
    assert _tail_line(records_path).startswith("#META")


def test_e2e6_corrupt_tail_with_corrupt_bak_raises_state_integrity_error(tmp_path: Path) -> None:
    config1 = _config(tmp_path, _SOURCE_LOG)
    pipeline.run(config1, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)
    config2 = Config(input_path=_BATCH_NEW, state_dir=config1.state_dir, out_dir=config1.out_dir)
    pipeline.run(config2, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE)  # creates .bak

    records_path = config1.state_dir / "records.csv"
    bak_path = config1.state_dir / "records.csv.bak"
    _corrupt_tail_sha(records_path)
    _corrupt_tail_sha(bak_path)

    with pytest.raises(StateIntegrityError):
        state.load(config1.state_dir)  # both copies unrecoverable -> fail loud, never silent


# --------------------------------------------------------------------
# E2E-7 -- docker/example repeatable demo scenario (design.md §4.7.7,
# §7.2, REQ3+REQ4). The committed docker/example/state/records.csv
# seed (16 rows, all BATCH_ID=1, byte-identical to
# expected_records_e2e1.csv) plus docker/example/input/week-2026-07-13.csv
# (3 brand-new-REQUEST_ID rows) together exercise all three WEEKLY
# ACCESS cases in one run: a brand-new IP (weekly == total), one
# overlapping IP (weekly < total), and nine seed-only IPs (weekly
# renders "-"). This is the exact scenario documented in
# usage.md's repeatable demo section.
# --------------------------------------------------------------------

_DOCKER_EXAMPLE_DIR: Path = _REPO_ROOT / "docker" / "example"
_DOCKER_EXAMPLE_SEED_STATE: Path = _DOCKER_EXAMPLE_DIR / "state" / "records.csv"
_DOCKER_EXAMPLE_WEEK_INPUT: Path = _DOCKER_EXAMPLE_DIR / "input" / "week-2026-07-13.csv"

#: Arbitrary (design.md: only affects the deliverable filename, never
#: WEEKLY/TOTAL semantics) -- a Monday after the 2026-07-13 input date.
_DOCKER_EXAMPLE_RUN_DATE: date = date(2026, 7, 15)

#: The 11-row 院所分析 table this scenario produces, first-seen order
#: (10 seed IPs, then the one brand-new IP last) -- verified by
#: simulation when this fixture pair was authored (REQ4 example_scenario).
_DOCKER_EXAMPLE_WEEKLY: tuple[object, ...] = ("-", "-", "-", "-", "-", "-", "-", 2, "-", "-", 1)
_DOCKER_EXAMPLE_TOTAL: tuple[int, ...] = (1, 1, 1, 1, 1, 1, 1, 9, 1, 1, 1)


def test_e2e7_docker_example_scenario_demonstrates_weekly_vs_total(tmp_path: Path) -> None:
    # Copy the committed seed state into a scratch state_dir -- the
    # checked-in docker/example/state/records.csv must stay pristine
    # (design.md §4.7.7, REQ4) so this test is repeatable run after run.
    state_dir = tmp_path / "state"
    state_dir.mkdir()
    (state_dir / "records.csv").write_bytes(_DOCKER_EXAMPLE_SEED_STATE.read_bytes())

    config = Config(
        input_path=_DOCKER_EXAMPLE_WEEK_INPUT, state_dir=state_dir, out_dir=tmp_path / "output"
    )
    summary = pipeline.run(
        config, run_date=_DOCKER_EXAMPLE_RUN_DATE, reference_path=_REAL_REFERENCE
    )

    assert summary.rows_in == 3
    assert summary.normal == 3
    assert summary.dropped_nonnormal == 0
    assert summary.new_records == 3
    assert summary.skipped_cross_state == 0
    assert summary.skipped_intra_batch == 0
    assert summary.state_total == 19
    assert summary.unique_ips == 11
    assert summary.batch_seq == 2
    assert summary.unmapped_hosp_ids == 0

    agg_sheet = load_workbook(summary.deliverable)["院所分析"]
    assert agg_sheet.max_row == 12  # header + 11 unique IPs
    rows_by_ip = {row[0].value: row for row in agg_sheet.iter_rows(min_row=2)}
    assert len(rows_by_ip) == 11

    weekly = [row[3].value for row in agg_sheet.iter_rows(min_row=2)]
    total = [row[4].value for row in agg_sheet.iter_rows(min_row=2)]
    assert weekly == list(_DOCKER_EXAMPLE_WEEKLY)
    assert total == list(_DOCKER_EXAMPLE_TOTAL)
    assert sum(v for v in weekly if isinstance(v, int)) == 3  # == this week's batch size
    assert sum(total) == 19  # == state_total

    # Brand-new IP this week -- WEEKLY ACCESS == TOTAL ACCESS (REQ3 case 1).
    brand_new = rows_by_ip["10.250.77.10"]
    assert brand_new[1].value == "3507030143"
    assert brand_new[2].value == "瀚田診所"
    assert brand_new[3].value == 1
    assert brand_new[4].value == 1

    # Overlap: present in both the seed AND this week -- WEEKLY < TOTAL
    # (REQ3 case 2).
    assert rows_by_ip["10.245.1.125"][3].value == 2
    assert rows_by_ip["10.245.1.125"][4].value == 9

    # Older-only (seed-only): no rows this week -> WEEKLY ACCESS "-"
    # (REQ3 case 3).
    assert rows_by_ip["10.243.129.44"][3].value == "-"
    assert rows_by_ip["10.243.129.44"][4].value == 1

    # REQ4 (design.md §3.7.3 per-run 黃底, §4.7.7): this run's 3 imported
    # batch-2 rows ALL carry the solid yellow highlight; every one of the
    # 16 pre-existing batch-1 seed rows carries NO fill. This is the exact
    # observation the docker/example quickstart tells the operator to
    # verify by eye, so the demo scenario asserts it explicitly (same
    # check as E2E-3/E2E-4).
    records_sheet = load_workbook(summary.deliverable)["調閱紀錄"]
    assert records_sheet.max_row == 1 + summary.state_total  # header + 19 rows
    assert _highlighted_rows(records_sheet) == [18, 19, 20]  # the 3 week batch-2 rows
    for row_idx in range(2, 18):  # all 16 seed batch-1 rows: no fill
        assert records_sheet.cell(row=row_idx, column=1).fill.patternType is None

    # The committed seed state itself must stay untouched by this run
    # (this test only ever wrote to the tmp_path copy).
    assert _DOCKER_EXAMPLE_SEED_STATE.read_bytes() == _EXPECTED_RECORDS_E2E1.read_bytes()


# --------------------------------------------------------------------
# Invariant: transform(source-log NORMAL) == 16 expected records
# (design.md §7.2 "不變量", §1.5.1). Pure-function level, independent of
# state/dedup/aggregate/xlsx -- reuses the SAME checked-in golden master
# as E2E-1 (see _load_expected_transformed_records's docstring).
# --------------------------------------------------------------------


def test_invariant_transform_matches_expected_snapshot() -> None:
    hosp_table = lookup.load(_REAL_REFERENCE)
    numbered_rows, unknown_status_skipped = csv_reader.read(_SOURCE_LOG)
    assert unknown_status_skipped == 0

    normal_rows = transform.filter_normal(numbered_rows)
    assert len(normal_rows) == 16

    transformed = transform.project(normal_rows, hosp_table)

    expected = _load_expected_transformed_records()
    assert len(expected) == 16
    assert transformed == expected


# --------------------------------------------------------------------
# Supplementary fixture-driven edge cases (design.md §6; exercises the
# remaining checked-in fixtures: empty.csv, all_nonnormal.csv,
# status_mixed_case.csv, hosp_map_small.csv).
# --------------------------------------------------------------------


def test_empty_input_against_empty_state_yields_zero_row_deliverable(tmp_path: Path) -> None:
    config = _config(tmp_path, _EMPTY)

    summary = pipeline.run(config, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)

    assert summary.rows_in == 0
    assert summary.new_records == 0
    assert summary.state_total == 0
    workbook = load_workbook(summary.deliverable)
    assert workbook["調閱紀錄"].max_row == 1  # header only -- 0 data rows, 0 yellow
    assert workbook["院所分析"].max_row == 1


def test_all_nonnormal_batch_leaves_existing_state_and_highlight_untouched(
    tmp_path: Path,
) -> None:
    config1 = _config(tmp_path, _SOURCE_LOG)
    pipeline.run(config1, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)

    config2 = Config(
        input_path=_ALL_NONNORMAL, state_dir=config1.state_dir, out_dir=config1.out_dir
    )
    summary = pipeline.run(config2, run_date=_RUN_DATE_2, reference_path=_REAL_REFERENCE)

    assert summary.normal == 0
    assert summary.new_records == 0
    assert summary.state_total == 16  # unchanged
    assert summary.batch_seq == 1  # still the pre-existing latest batch

    records_sheet = load_workbook(summary.deliverable)["調閱紀錄"]
    assert _highlighted_rows(records_sheet) == list(range(2, 18))  # unchanged: all 16, batch 1


def test_status_mixed_case_is_all_treated_as_normal(tmp_path: Path) -> None:
    config = _config(tmp_path, _STATUS_MIXED_CASE)

    summary = pipeline.run(config, run_date=_RUN_DATE_1, reference_path=_REAL_REFERENCE)

    assert summary.normal == 2  # "NORMAL" / "normal" all counted (Excel `=` semantics)
    assert summary.new_records == 2
    assert summary.state_total == 2
    result = state.load(config.state_dir)
    assert {r.batch_id for r in result.existing} == {1}


def test_unmapped_hosp_abbr_reads_back_as_none(tmp_path: Path) -> None:
    small_reference = _gzip_reference(_HOSP_MAP_SMALL, tmp_path)
    config = _config(tmp_path, _SOURCE_LOG)

    summary = pipeline.run(config, run_date=_RUN_DATE_1, reference_path=small_reference)

    # hosp_map_small.csv includes 0937010019 but deliberately omits
    # 1503190020 -- exactly one of the 16 NORMAL rows' HOSP_ID is unmapped.
    assert summary.unmapped_hosp_ids == 1
    assert summary.unique_ips == 10  # aggregation is unaffected by the lookup table

    workbook = load_workbook(summary.deliverable)
    agg_sheet = workbook["院所分析"]
    shuchuan = next(row for row in agg_sheet.iter_rows(min_row=2) if row[1].value == "0937010019")
    assert shuchuan[2].value == "秀傳醫院"  # still resolves -- included in the small map

    unmapped_agg_row = next(
        row for row in agg_sheet.iter_rows(min_row=2) if row[0].value == "10.238.23.241"
    )
    assert unmapped_agg_row[1].value == "1503190020"
    # design.md §1.5.9/§3.7.6: a "" HOSP_ABBR reads back as None, not "".
    assert unmapped_agg_row[2].value is None

    records_sheet = workbook["調閱紀錄"]
    unmapped_state_row = next(
        row
        for row in records_sheet.iter_rows(min_row=2)
        if row[2].value == "10.238.23.241"  # column C = CLIENT IP
    )
    assert unmapped_state_row[5].value is None  # column F = HOSP_ABBR
