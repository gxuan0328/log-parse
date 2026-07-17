"""Unit tests for report_export.xlsx_writer (design.md §7.1
test_xlsx_writer, §3.7, §3.7.6 fidelity read-back)."""

from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from report_export import xlsx_writer
from report_export.errors import WriteError
from report_export.models import ReportRow, StateRecord


def _record(
    batch_id: int = 1,
    request_id: str = "req-1",
    *,
    app_time_iso: str = "2026-07-05 16:03:34.359",
    client_ip: str = "10.245.1.125",
    hosp_id: str = "0937010019",
    hosp_abbr: str = "秀傳醫院",
) -> StateRecord:
    return StateRecord(
        batch_id=batch_id,
        request_id=request_id,
        app_time_iso=app_time_iso,
        client_ip=client_ip,
        server_ip="10.21.3.35",
        hosp_id=hosp_id,
        hosp_abbr=hosp_abbr,
        prsn_id="21B026FA29B9A716C23AEB10D3F04A63",
        birthday="19560711",
        patient_id_aes="2EDEBACB75D9FA547F2018E13E695AF1",
    )


def _report_row(
    client_ip: str = "10.245.1.125",
    hosp_id: str = "0937010019",
    hosp_abbr: str = "秀傳醫院",
    weekly_access: int = 1,
    total_access: int = 1,
) -> ReportRow:
    return ReportRow(
        client_ip=client_ip,
        hosp_id=hosp_id,
        hosp_abbr=hosp_abbr,
        weekly_access=weekly_access,
        total_access=total_access,
    )


def _save_and_reload(workbook: Workbook, tmp_path: Path, name: str = "out.xlsx") -> Workbook:
    dest = tmp_path / name
    workbook.save(str(dest))
    return load_workbook(dest)


# --------------------------------------------------------------------
# Workbook shape: exactly 2 sheets, zero formulas, exact headers
# (design.md §3.7.1, §3.7.6)
# --------------------------------------------------------------------


def test_workbook_has_exactly_two_sheets_in_order() -> None:
    workbook = xlsx_writer.build_workbook([_record()], [_report_row()])
    assert workbook.sheetnames == ["調閱紀錄", "院所分析"]


def test_workbook_default_sheet_is_removed() -> None:
    workbook = xlsx_writer.build_workbook([], [])
    assert "Sheet" not in workbook.sheetnames


def test_records_sheet_header_is_exact() -> None:
    workbook = xlsx_writer.build_workbook([], [])
    sheet = workbook["調閱紀錄"]
    header = [cell.value for cell in sheet[1]]
    assert header == [
        "DATE",
        "TIME",
        "CLIENT IP",
        "SERVER IP",
        "HOSP_ID",
        "HOSP_ABBR",
        "PRSN_ID",
        "BIRTHDAY",
        "PATIENT ID AES",
    ]


def test_aggregate_sheet_header_is_exact() -> None:
    workbook = xlsx_writer.build_workbook([], [])
    sheet = workbook["院所分析"]
    header = [cell.value for cell in sheet[1]]
    assert header == ["CLIENT IP", "HOSP_ID", "HOSP_ABBR", "WEEKLY ACCESS", "TOTAL ACCESS"]


def test_no_cell_holds_a_formula(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook([_record()], [_report_row()])
    reloaded = _save_and_reload(workbook, tmp_path)
    for sheet in reloaded.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                assert cell.data_type != "f"
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("=")


def test_workbook_contains_no_reference_sheets() -> None:
    workbook = xlsx_writer.build_workbook([_record()], [_report_row()])
    for forbidden in ("紀錄匯入", "格式轉換", "HOSP_ID對照表"):
        assert forbidden not in workbook.sheetnames


# --------------------------------------------------------------------
# §3.7.6 fidelity read-back: A2==B2 datetime, number_format, TEXT types
# --------------------------------------------------------------------


def test_date_and_time_share_the_same_datetime_value_after_round_trip(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook(
        [_record(app_time_iso="2026-07-05 16:03:34.359")], []
    )
    reloaded = _save_and_reload(workbook, tmp_path)
    sheet = reloaded["調閱紀錄"]
    assert sheet["A2"].value == datetime(2026, 7, 5, 16, 3, 34, 359000)
    assert sheet["A2"].value == sheet["B2"].value


def test_date_cell_number_format_contains_yyyy(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook([_record()], [])
    reloaded = _save_and_reload(workbook, tmp_path)
    assert "yyyy" in reloaded["調閱紀錄"]["A2"].number_format


def test_time_cell_number_format_contains_hmm(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook([_record()], [])
    reloaded = _save_and_reload(workbook, tmp_path)
    assert "h:mm" in reloaded["調閱紀錄"]["B2"].number_format


def test_no_millisecond_precision_is_lost_without_fractional_seconds(tmp_path: Path) -> None:
    # design.md §3.8 S4: no-ms APP_TIME is an accepted, compatible format.
    workbook = xlsx_writer.build_workbook(
        [_record(app_time_iso="2026-07-05 16:03:34")], []
    )
    reloaded = _save_and_reload(workbook, tmp_path)
    assert reloaded["調閱紀錄"]["A2"].value == datetime(2026, 7, 5, 16, 3, 34)


def test_hosp_id_round_trips_as_string_with_leading_zero(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook([_record(hosp_id="0937010019")], [])
    reloaded = _save_and_reload(workbook, tmp_path)
    cell = reloaded["調閱紀錄"]["E2"]
    assert cell.value == "0937010019"
    assert isinstance(cell.value, str)


def test_unmapped_hosp_abbr_round_trips_as_empty_or_none(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook([_record(hosp_abbr="")], [])
    reloaded = _save_and_reload(workbook, tmp_path)
    value = reloaded["調閱紀錄"]["F2"].value
    assert value in ("", None)


def test_text_columns_get_at_number_format(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook([_record()], [])
    reloaded = _save_and_reload(workbook, tmp_path)
    sheet = reloaded["調閱紀錄"]
    for column in "CDEFGHI":
        assert sheet[f"{column}2"].number_format == "@"


def test_aggregate_sheet_text_columns_round_trip_as_strings(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook(
        [], [_report_row(client_ip="10.1.1.1", hosp_id="0937010019", hosp_abbr="秀傳醫院")]
    )
    reloaded = _save_and_reload(workbook, tmp_path)
    sheet = reloaded["院所分析"]
    assert sheet["A2"].value == "10.1.1.1"
    assert sheet["B2"].value == "0937010019"
    assert isinstance(sheet["B2"].value, str)
    assert sheet["C2"].value == "秀傳醫院"


def test_aggregate_unmapped_hosp_abbr_round_trips_as_empty_or_none(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook([], [_report_row(hosp_abbr="")])
    reloaded = _save_and_reload(workbook, tmp_path)
    assert reloaded["院所分析"]["C2"].value in ("", None)


def test_weekly_and_total_access_columns_round_trip_as_int(tmp_path: Path) -> None:
    workbook = xlsx_writer.build_workbook([], [_report_row(weekly_access=2, total_access=9)])
    reloaded = _save_and_reload(workbook, tmp_path)
    sheet = reloaded["院所分析"]
    weekly_cell = sheet["D2"]
    total_cell = sheet["E2"]
    assert weekly_cell.value == 2
    assert isinstance(weekly_cell.value, int)
    assert weekly_cell.number_format == "General"
    assert total_cell.value == 9
    assert isinstance(total_cell.value, int)
    assert total_cell.number_format == "General"


def test_weekly_access_zero_renders_as_dash_string(tmp_path: Path) -> None:
    # design.md §3.1.3 REQ3b: "本周無存取紀錄之院所則填入 -" -- the MODEL stays
    # numeric (weekly_access=0), but the rendered cell is the "-" string.
    workbook = xlsx_writer.build_workbook([], [_report_row(weekly_access=0, total_access=1)])
    reloaded = _save_and_reload(workbook, tmp_path)
    cell = reloaded["院所分析"]["D2"]
    assert cell.value == "-"
    assert isinstance(cell.value, str)
    assert cell.number_format == "@"
    # TOTAL ACCESS is unaffected by WEEKLY ACCESS being "-".
    assert reloaded["院所分析"]["E2"].value == 1


# --------------------------------------------------------------------
# Row ordering / content preserved verbatim (design.md §3.7.3, §3.1.3)
# --------------------------------------------------------------------


def test_records_sheet_preserves_full_state_order() -> None:
    records = [_record(1, "a", client_ip="10.0.0.1"), _record(1, "b", client_ip="10.0.0.2")]
    workbook = xlsx_writer.build_workbook(records, [])
    sheet = workbook["調閱紀錄"]
    assert sheet["C2"].value == "10.0.0.1"
    assert sheet["C3"].value == "10.0.0.2"


def test_aggregate_sheet_preserves_report_row_order() -> None:
    rows = [_report_row(client_ip="10.0.0.2"), _report_row(client_ip="10.0.0.1")]
    workbook = xlsx_writer.build_workbook([], rows)
    sheet = workbook["院所分析"]
    assert sheet["A2"].value == "10.0.0.2"
    assert sheet["A3"].value == "10.0.0.1"


def test_empty_state_and_report_rows_produce_header_only_sheets() -> None:
    workbook = xlsx_writer.build_workbook([], [])
    assert workbook["調閱紀錄"].max_row == 1
    assert workbook["院所分析"].max_row == 1


# --------------------------------------------------------------------
# Yellow highlight -- latest batch only (design.md §3.7.3, §1.5.9)
# --------------------------------------------------------------------


def test_latest_batch_rows_get_yellow_fill() -> None:
    records = [_record(batch_id=1, request_id="a"), _record(batch_id=2, request_id="b")]
    workbook = xlsx_writer.build_workbook(records, [])
    sheet = workbook["調閱紀錄"]
    for column in "ABCDEFGHI":
        cell = sheet[f"{column}3"]  # row 2 of data = batch 2 = latest
        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb.endswith("FFFF00")


def test_older_batch_rows_get_no_fill() -> None:
    records = [_record(batch_id=1, request_id="a"), _record(batch_id=2, request_id="b")]
    workbook = xlsx_writer.build_workbook(records, [])
    sheet = workbook["調閱紀錄"]
    for column in "ABCDEFGHI":
        assert sheet[f"{column}2"].fill.patternType is None  # row 2 = batch 1 = not latest


def test_all_rows_highlighted_when_single_batch() -> None:
    records = [_record(batch_id=1, request_id="a"), _record(batch_id=1, request_id="b")]
    workbook = xlsx_writer.build_workbook(records, [])
    sheet = workbook["調閱紀錄"]
    for row_idx in (2, 3):
        assert sheet[f"A{row_idx}"].fill.patternType == "solid"


def test_highlight_survives_save_and_reload(tmp_path: Path) -> None:
    records = [_record(batch_id=1, request_id="a"), _record(batch_id=2, request_id="b")]
    workbook = xlsx_writer.build_workbook(records, [])
    reloaded = _save_and_reload(workbook, tmp_path)
    sheet = reloaded["調閱紀錄"]
    assert sheet["A3"].fill.fgColor.rgb == "FFFFFF00"
    assert sheet["A2"].fill.patternType is None


def test_aggregate_sheet_rows_never_get_yellow_fill() -> None:
    workbook = xlsx_writer.build_workbook([], [_report_row()])
    sheet = workbook["院所分析"]
    for column in "ABCDE":
        assert sheet[f"{column}2"].fill.patternType is None


def test_empty_full_state_produces_no_highlighted_rows() -> None:
    workbook = xlsx_writer.build_workbook([], [])
    # No data rows exist at all -- nothing to assert per-cell; the
    # sheet must simply not error out building max()'s default=0 path.
    assert workbook["調閱紀錄"].max_row == 1


# --------------------------------------------------------------------
# Header style (design.md §3.7.5)
# --------------------------------------------------------------------


def test_header_cells_are_bold_size_12() -> None:
    workbook = xlsx_writer.build_workbook([], [])
    cell = workbook["調閱紀錄"]["A1"]
    assert cell.font.bold is True
    assert cell.font.size == 12


def test_header_cells_have_explicit_green_fill() -> None:
    workbook = xlsx_writer.build_workbook([], [])
    cell = workbook["調閱紀錄"]["A1"]
    assert cell.fill.patternType == "solid"
    # 8-hex with explicit FF alpha (design.md §1.5.9 gotcha): a bare
    # 6-hex 'E2EFDA' round-trips as '00E2EFDA' (alpha=00, invisible).
    assert cell.fill.fgColor.rgb == "FFE2EFDA"


def test_both_sheet_headers_are_styled() -> None:
    workbook = xlsx_writer.build_workbook([], [])
    cell = workbook["院所分析"]["A1"]
    assert cell.font.bold is True
    assert cell.fill.patternType == "solid"


# --------------------------------------------------------------------
# REQ1 formatting: center alignment, thin data borders, thick header
# bottom border, auto-fit column widths (design.md §3.7.3-§3.7.5, §3.7.6)
# --------------------------------------------------------------------


def test_every_data_cell_is_centered_with_thin_border_on_both_sheets() -> None:
    workbook = xlsx_writer.build_workbook([_record()], [_report_row()])
    for sheet_name in ("調閱紀錄", "院所分析"):
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                assert cell.alignment.horizontal == "center"
                assert cell.alignment.vertical == "center"
                assert cell.border.left.style == "thin"
                assert cell.border.right.style == "thin"
                assert cell.border.top.style == "thin"
                assert cell.border.bottom.style == "thin"


def test_every_header_cell_is_centered_with_thick_bottom_border_on_both_sheets() -> None:
    workbook = xlsx_writer.build_workbook([_record()], [_report_row()])
    for sheet_name in ("調閱紀錄", "院所分析"):
        sheet = workbook[sheet_name]
        for cell in sheet[1]:
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
            assert cell.border.bottom.style == "thick"
            # design.md §3.7.5: thin sides/top are cosmetic grid
            # continuity with the data rows below -- the REQUIRED
            # emphasis is the thick bottom border.
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.top.style == "thin"


def test_yellow_fill_and_thin_border_and_center_coexist_on_latest_batch_rows() -> None:
    records = [_record(batch_id=1, request_id="a"), _record(batch_id=2, request_id="b")]
    workbook = xlsx_writer.build_workbook(records, [])
    sheet = workbook["調閱紀錄"]
    for column in "ABCDEFGHI":
        cell = sheet[f"{column}3"]  # latest (highlighted) batch-2 row
        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb == "FFFFFF00"
        assert cell.alignment.horizontal == "center"
        assert cell.border.left.style == "thin"
        assert cell.border.bottom.style == "thin"


def test_column_widths_match_e2e1_anchors_on_a_representative_state() -> None:
    # design.md §3.7.3/§3.7.4 REQ1d: width = round(max(display_width(header),
    # max(display_width(rendered data))) * 1.2, 2). A SINGLE record/row
    # whose fields already hit the same per-column maxima as the real
    # 16-row/10-IP E2E-1 anchor state (design.md §1.5.1/§1.5.2) reproduces
    # the EXACT same widths, since max() over a superset can never
    # exceed max() over a subset that already holds the maximal value.
    record = _record(
        client_ip="10.241.189.173",  # longest CLIENT_IP/SERVER_IP-column IP (14 chars)
        hosp_abbr="彰基二林醫",  # widest HOSP_ABBR in the anchor dataset (5 CJK = 10)
    )
    report_row = _report_row(
        client_ip="10.241.189.173", hosp_abbr="彰基二林醫", weekly_access=1, total_access=1
    )
    workbook = xlsx_writer.build_workbook([record], [report_row])

    records_sheet = workbook["調閱紀錄"]
    expected_records_widths = {
        "A": 12.0,  # DATE: "YYYY-MM-DD" (10) * 1.2
        "B": 9.6,  # TIME: "HH:MM:SS" (8) * 1.2
        "C": 16.8,  # CLIENT IP: "10.241.189.173" (14) * 1.2
        "D": 12.0,  # SERVER IP: "10.21.3.35" (10) * 1.2
        "E": 12.0,  # HOSP_ID: 10-digit string * 1.2
        "F": 12.0,  # HOSP_ABBR: "彰基二林醫" (width 10) * 1.2
        "G": 38.4,  # PRSN_ID: hex32 (32) * 1.2
        "H": 9.6,  # BIRTHDAY: "19560711" (8) * 1.2
        "I": 38.4,  # PATIENT ID AES: hex32 (32) * 1.2
    }
    for column, width in expected_records_widths.items():
        assert records_sheet.column_dimensions[column].width == width

    agg_sheet = workbook["院所分析"]
    expected_agg_widths = {
        "A": 16.8,  # CLIENT IP: same 14-char IP * 1.2
        "B": 12.0,  # HOSP_ID: 10-digit string * 1.2
        "C": 12.0,  # HOSP_ABBR: width 10 * 1.2
        "D": 15.6,  # WEEKLY ACCESS: header "WEEKLY ACCESS" (13) * 1.2 (dominates 1-digit data)
        "E": 14.4,  # TOTAL ACCESS: header "TOTAL ACCESS" (12) * 1.2
    }
    for column, width in expected_agg_widths.items():
        assert agg_sheet.column_dimensions[column].width == width


def test_column_widths_are_header_based_on_empty_sheets() -> None:
    # design.md §3.7.3 REQ1d: a header-only (0 data row) sheet still
    # gets a sensible, header-derived width -- never a crash, never 0.
    workbook = xlsx_writer.build_workbook([], [])
    records_sheet = workbook["調閱紀錄"]
    assert records_sheet.column_dimensions["A"].width == round(len("DATE") * 1.2, 2)
    agg_sheet = workbook["院所分析"]
    assert agg_sheet.column_dimensions["D"].width == round(len("WEEKLY ACCESS") * 1.2, 2)


# --------------------------------------------------------------------
# resolve_filename() -- same-day disambiguation (design.md §3.7.2)
# --------------------------------------------------------------------


def test_resolve_filename_uses_base_name_when_no_existing_file(tmp_path: Path) -> None:
    name = xlsx_writer.resolve_filename(
        run_date=date(2026, 7, 15), out_dir=tmp_path, today_runs=[], input_sha256="abc"
    )
    assert name == "2026-07-15_連線紀錄.xlsx"


def test_resolve_filename_same_sha_reuses_base_name(tmp_path: Path) -> None:
    (tmp_path / "2026-07-15_連線紀錄.xlsx").write_bytes(b"")
    today_runs = [{"input_sha256": "abc"}]
    name = xlsx_writer.resolve_filename(
        run_date=date(2026, 7, 15), out_dir=tmp_path, today_runs=today_runs, input_sha256="abc"
    )
    assert name == "2026-07-15_連線紀錄.xlsx"


def test_resolve_filename_different_sha_appends_suffix(tmp_path: Path) -> None:
    (tmp_path / "2026-07-15_連線紀錄.xlsx").write_bytes(b"")
    today_runs = [{"input_sha256": "abc"}]
    name = xlsx_writer.resolve_filename(
        run_date=date(2026, 7, 15),
        out_dir=tmp_path,
        today_runs=today_runs,
        input_sha256="different",
    )
    assert name == "2026-07-15_連線紀錄_02.xlsx"


def test_resolve_filename_third_distinct_batch_gets_seq_03(tmp_path: Path) -> None:
    (tmp_path / "2026-07-15_連線紀錄.xlsx").write_bytes(b"")
    today_runs = [{"input_sha256": "abc"}, {"input_sha256": "def"}]
    name = xlsx_writer.resolve_filename(
        run_date=date(2026, 7, 15), out_dir=tmp_path, today_runs=today_runs, input_sha256="ghi"
    )
    assert name == "2026-07-15_連線紀錄_03.xlsx"


def test_resolve_filename_existing_file_but_no_runs_logged_disambiguates_safely(
    tmp_path: Path,
) -> None:
    # Defensive edge case: a file exists but runs.jsonl has no record
    # for today at all (e.g. a lost/reset runs.jsonl). Never silently
    # reuse/clobber it without a confirmed sha256 match -- the
    # disambiguated suffix starts at 2 (the existing file is presumed
    # to already be batch 1 of the day).
    (tmp_path / "2026-07-15_連線紀錄.xlsx").write_bytes(b"")
    name = xlsx_writer.resolve_filename(
        run_date=date(2026, 7, 15), out_dir=tmp_path, today_runs=[], input_sha256="abc"
    )
    assert name == "2026-07-15_連線紀錄_02.xlsx"


# --------------------------------------------------------------------
# write() -- tmp write + fsync, no rename (design.md §3.8 S8)
# --------------------------------------------------------------------


def test_write_creates_tmp_file_not_final_file(tmp_path: Path) -> None:
    tmp_file, final_file = xlsx_writer.write(
        tmp_path,
        [_record()],
        [_report_row()],
        run_date=date(2026, 7, 15),
        today_runs=[],
        input_sha256="abc",
    )
    assert tmp_file.exists()
    assert not final_file.exists()
    assert tmp_file.name == "2026-07-15_連線紀錄.xlsx.tmp"
    assert final_file.name == "2026-07-15_連線紀錄.xlsx"


def test_write_tmp_file_is_a_valid_readable_workbook(tmp_path: Path) -> None:
    tmp_file, _final_file = xlsx_writer.write(
        tmp_path,
        [_record()],
        [_report_row()],
        run_date=date(2026, 7, 15),
        today_runs=[],
        input_sha256="abc",
    )
    # openpyxl.load_workbook() gates on a `.xlsx`-family extension when
    # given a path, so a `.tmp`-suffixed path must be opened as a
    # file-like object instead (design.md §3.8 S8: the tmp file is a
    # genuine, fully-written xlsx -- only its name differs).
    with tmp_file.open("rb") as fh:
        reloaded = load_workbook(fh)
    assert reloaded.sheetnames == ["調閱紀錄", "院所分析"]


def test_write_creates_out_dir_if_missing(tmp_path: Path) -> None:
    out_dir = tmp_path / "brand" / "new"
    tmp_file, _final_file = xlsx_writer.write(
        out_dir, [], [], run_date=date(2026, 7, 15), today_runs=[], input_sha256="abc"
    )
    assert tmp_file.exists()


def test_write_out_dir_gets_0700_permissions(tmp_path: Path) -> None:
    out_dir = tmp_path / "fresh_out"
    xlsx_writer.write(
        out_dir, [], [], run_date=date(2026, 7, 15), today_runs=[], input_sha256="abc"
    )
    assert oct(out_dir.stat().st_mode & 0o777) == "0o700"


def test_write_tmp_file_gets_0600_permissions(tmp_path: Path) -> None:
    tmp_file, _final_file = xlsx_writer.write(
        tmp_path, [], [], run_date=date(2026, 7, 15), today_runs=[], input_sha256="abc"
    )
    assert oct(tmp_file.stat().st_mode & 0o777) == "0o600"


def test_write_wraps_unexpected_save_failure_as_write_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    real_chmod = os.chmod

    def _selective_raise(
        path: int | str | bytes | os.PathLike[str] | os.PathLike[bytes], mode: int
    ) -> None:
        # Only fail the deliverable tmp file's own chmod (0o600) --
        # `_ensure_out_dir`'s directory chmod (0o700) must still
        # succeed, so this isolates the write()-body except branch.
        if mode == 0o600:
            raise OSError(5, "Input/output error")
        real_chmod(path, mode)

    monkeypatch.setattr(os, "chmod", _selective_raise)
    with pytest.raises(WriteError, match="cannot write deliverable workbook"):
        xlsx_writer.write(
            tmp_path, [], [], run_date=date(2026, 7, 15), today_runs=[], input_sha256="abc"
        )


def test_ensure_out_dir_wraps_unexpected_mkdir_failure_as_write_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    def _raise_eio(
        self: Path, *, mode: int = 0o777, parents: bool = False, exist_ok: bool = False
    ) -> None:
        raise OSError(5, "Input/output error")

    monkeypatch.setattr(Path, "mkdir", _raise_eio)
    with pytest.raises(WriteError, match="cannot prepare out_dir"):
        xlsx_writer.write(
            tmp_path / "brand" / "new",
            [],
            [],
            run_date=date(2026, 7, 15),
            today_runs=[],
            input_sha256="abc",
        )


# --------------------------------------------------------------------
# _parse_app_time_iso fail-loud guard (design.md: never silently
# mis-render a corrupted records.csv, §4.7.7 R7)
# --------------------------------------------------------------------


def test_unparsable_app_time_iso_raises_write_error() -> None:
    bad_record = _record(app_time_iso="not-a-valid-timestamp")
    with pytest.raises(WriteError, match="cannot render APP_TIME_ISO"):
        xlsx_writer.build_workbook([bad_record], [])


def test_write_respects_same_day_disambiguation(tmp_path: Path) -> None:
    (tmp_path / "2026-07-15_連線紀錄.xlsx").write_bytes(b"")
    _tmp_file, final_file = xlsx_writer.write(
        tmp_path,
        [],
        [],
        run_date=date(2026, 7, 15),
        today_runs=[{"input_sha256": "abc"}],
        input_sha256="different-batch",
    )
    assert final_file.name == "2026-07-15_連線紀錄_02.xlsx"
