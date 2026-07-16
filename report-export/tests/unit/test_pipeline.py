"""Unit tests for report_export.pipeline (design.md §7.1 test_pipeline, §3.8, §3.9.2)."""

from __future__ import annotations

import csv
import gzip
import json
import os
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from report_export import pipeline, state, statelock
from report_export.config import Config
from report_export.errors import (
    InputValidationError,
    LockBusyError,
    WriteError,
)
from report_export.errors import (
    ReferenceError as ReportExportReferenceError,
)

_HEADER: tuple[str, ...] = (
    "REGION",
    "STATUS",
    "API_TIME",
    "APP_TIME",
    "DELTA_SEC",
    "VERIFY_STATUS",
    "REQUEST_ID",
    "API_SERVER",
    "APP_SERVER",
    "HOSP_ID",
    "PRSN_ID",
    "CLIENT_IP",
    "PATIENT_ID_AES",
    "BIRTHDAY",
)


def _row(
    request_id: str,
    *,
    status: str = "NORMAL",
    app_time: str = "2026-07-05 16:03:34.359",
    hosp_id: str = "0937010019",
    client_ip: str = "10.245.1.125",
) -> tuple[str, ...]:
    return (
        "台北",
        status,
        "2026-07-05 16:03:24.381",
        app_time,
        "9.978",
        "OK",
        request_id,
        "10.22.63.37",
        "10.21.3.35",
        hosp_id,
        "21B026FA29B9A716C23AEB10D3F04A63",
        client_ip,
        "2EDEBACB75D9FA547F2018E13E695AF1",
        "19560711",
    )


def _write_input_csv(path: Path, rows: list[tuple[str, ...]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(_HEADER)
        writer.writerows(rows)


def _write_reference(path: Path, entries: dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(path, mode="wt", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(("HOSP_ID", "HOSP_ABBR"))
        for hosp_id, hosp_abbr in entries.items():
            writer.writerow((hosp_id, hosp_abbr))


@pytest.fixture
def reference_path(tmp_path: Path) -> Path:
    path = tmp_path / "reference" / "hosp_id_map.csv.gz"
    _write_reference(path, {"0937010019": "秀傳醫院", "1145010038": "門諾醫院"})
    return path


def _config(tmp_path: Path, input_path: Path) -> Config:
    return Config(input_path=input_path, state_dir=tmp_path / "state", out_dir=tmp_path / "output")


# --------------------------------------------------------------------
# First run -- empty state (design.md §3.5.5, §6-10)
# --------------------------------------------------------------------


def test_first_run_creates_batch_1(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1"), _row("req-2")])
    config = _config(tmp_path, input_path)

    summary = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert summary.batch_seq == 1
    assert summary.new_records == 2
    assert summary.state_total == 2
    result = state.load(config.state_dir)
    assert {r.batch_id for r in result.existing} == {1}


def test_first_run_summary_fields(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(
        input_path, [_row("req-1"), _row("req-2"), _row("req-orphan", status="ORPHAN")]
    )
    config = _config(tmp_path, input_path)

    summary = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert summary.rows_in == 3
    assert summary.normal == 2
    assert summary.dropped_nonnormal == 1
    assert summary.unknown_status_skipped == 0
    assert summary.skipped_cross_state == 0
    assert summary.skipped_intra_batch == 0
    assert summary.unmapped_hosp_ids == 0
    assert summary.unique_ips == 1
    assert summary.run_date == "2026-07-15"
    assert summary.input == str(input_path)
    assert len(summary.input_sha256) == 64  # hex sha256


def test_unmapped_hosp_id_is_counted_and_does_not_fail(
    tmp_path: Path, reference_path: Path
) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1", hosp_id="9999999999")])
    config = _config(tmp_path, input_path)

    summary = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert summary.unmapped_hosp_ids == 1
    assert summary.new_records == 1


def test_deliverable_file_is_created(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)

    summary = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    deliverable = Path(summary.deliverable)
    assert deliverable.exists()
    assert deliverable.name == "2026-07-15_連線紀錄.xlsx"
    assert not deliverable.with_name(deliverable.name + ".tmp").exists()


def test_deliverable_reflects_anchor_data(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1"), _row("req-2")])
    config = _config(tmp_path, input_path)

    summary = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    workbook = load_workbook(summary.deliverable)
    assert workbook.sheetnames == ["調閱紀錄", "院所分析"]
    records_sheet = workbook["調閱紀錄"]
    assert records_sheet.max_row == 3  # header + 2 rows
    agg_sheet = workbook["院所分析"]
    assert agg_sheet["D2"].value == 2  # both rows share the same CLIENT_IP


# --------------------------------------------------------------------
# Idempotent rerun -- 0 new records (design.md §4.1, §6-4)
# --------------------------------------------------------------------


def test_rerun_same_input_yields_zero_new_records(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1"), _row("req-2")])
    config = _config(tmp_path, input_path)
    pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    summary = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert summary.new_records == 0
    assert summary.skipped_cross_state == 2
    assert summary.state_total == 2


def test_rerun_does_not_touch_records_csv_bytes(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)
    pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)
    records_path = config.state_dir / "records.csv"
    before = records_path.read_bytes()

    pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert records_path.read_bytes() == before
    assert not (config.state_dir / "records.csv.bak").exists()  # commit() never ran again


def test_rerun_still_produces_deliverable_reflecting_state(
    tmp_path: Path, reference_path: Path
) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)
    first = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    second = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert Path(second.deliverable).exists()
    # design.md §4.1: batch_seq always mirrors max(batch_id in
    # full_state) -- an idempotent 0-new rerun leaves that unchanged
    # (never reports a "next" batch number that no record carries).
    assert second.batch_seq == first.batch_seq == 1


def test_empty_input_produces_zero_new_records_and_deliverable(
    tmp_path: Path, reference_path: Path
) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [])
    config = _config(tmp_path, input_path)

    summary = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert summary.new_records == 0
    assert summary.state_total == 0
    assert Path(summary.deliverable).exists()


# --------------------------------------------------------------------
# Second distinct batch (design.md §3.8 S7, per-run yellow reset, §4.1)
# --------------------------------------------------------------------


def test_second_distinct_batch_gets_batch_id_2(tmp_path: Path, reference_path: Path) -> None:
    input_path_1 = tmp_path / "input1.csv"
    _write_input_csv(input_path_1, [_row("req-1")])
    config_1 = _config(tmp_path, input_path_1)
    pipeline.run(config_1, run_date=date(2026, 7, 15), reference_path=reference_path)

    input_path_2 = tmp_path / "input2.csv"
    _write_input_csv(input_path_2, [_row("req-2")])
    config_2 = Config(
        input_path=input_path_2, state_dir=config_1.state_dir, out_dir=config_1.out_dir
    )
    summary = pipeline.run(config_2, run_date=date(2026, 7, 16), reference_path=reference_path)

    assert summary.batch_seq == 2
    assert summary.state_total == 2
    result = state.load(config_1.state_dir)
    assert {r.batch_id for r in result.existing} == {1, 2}


def test_second_batch_deliverable_highlights_only_new_rows(
    tmp_path: Path, reference_path: Path
) -> None:
    input_path_1 = tmp_path / "input1.csv"
    _write_input_csv(input_path_1, [_row("req-1")])
    config_1 = _config(tmp_path, input_path_1)
    pipeline.run(config_1, run_date=date(2026, 7, 15), reference_path=reference_path)

    input_path_2 = tmp_path / "input2.csv"
    _write_input_csv(input_path_2, [_row("req-2")])
    config_2 = Config(
        input_path=input_path_2, state_dir=config_1.state_dir, out_dir=config_1.out_dir
    )
    summary = pipeline.run(config_2, run_date=date(2026, 7, 16), reference_path=reference_path)

    workbook = load_workbook(summary.deliverable)
    sheet = workbook["調閱紀錄"]
    assert sheet["A2"].fill.patternType is None  # batch 1, not latest
    assert sheet["A3"].fill.patternType == "solid"  # batch 2, latest


# --------------------------------------------------------------------
# run_date injection determines filename (design.md §3.9.1, §3.7.2)
# --------------------------------------------------------------------


def test_run_date_injection_determines_filename(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)

    summary = pipeline.run(config, run_date=date(2099, 1, 31), reference_path=reference_path)

    assert Path(summary.deliverable).name == "2099-01-31_連線紀錄.xlsx"


def test_default_run_date_is_today_when_not_injected(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)

    summary = pipeline.run(config, reference_path=reference_path)

    assert summary.run_date == date.today().strftime("%Y-%m-%d")


# --------------------------------------------------------------------
# Same-day distinct-batch filename disambiguation (design.md §3.7.2)
# --------------------------------------------------------------------


def test_same_day_distinct_batches_get_suffixed_filename(
    tmp_path: Path, reference_path: Path
) -> None:
    input_path_1 = tmp_path / "input1.csv"
    _write_input_csv(input_path_1, [_row("req-1")])
    config_1 = _config(tmp_path, input_path_1)
    first = pipeline.run(config_1, run_date=date(2026, 7, 15), reference_path=reference_path)

    input_path_2 = tmp_path / "input2.csv"
    _write_input_csv(input_path_2, [_row("req-2")])
    config_2 = Config(
        input_path=input_path_2, state_dir=config_1.state_dir, out_dir=config_1.out_dir
    )
    second = pipeline.run(config_2, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert Path(first.deliverable).name == "2026-07-15_連線紀錄.xlsx"
    assert Path(second.deliverable).name == "2026-07-15_連線紀錄_02.xlsx"


def test_same_day_same_input_rerun_reuses_filename(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)
    first = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    second = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    assert first.deliverable == second.deliverable


# --------------------------------------------------------------------
# Errors propagate untouched (design.md §4.2 -- cli.py is the sole catcher)
# --------------------------------------------------------------------


def test_missing_reference_raises_reference_error(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)

    with pytest.raises(ReportExportReferenceError):
        pipeline.run(
            config, run_date=date(2026, 7, 15), reference_path=tmp_path / "no-such-ref.csv.gz"
        )


def test_malformed_input_raises_input_validation_error(
    tmp_path: Path, reference_path: Path
) -> None:
    input_path = tmp_path / "input.csv"
    input_path.write_text("NOT,THE,RIGHT,HEADER\n", encoding="utf-8")
    config = _config(tmp_path, input_path)

    with pytest.raises(InputValidationError):
        pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)


def test_missing_input_raises_input_validation_error(tmp_path: Path, reference_path: Path) -> None:
    config = _config(tmp_path, tmp_path / "does-not-exist.csv")

    with pytest.raises(InputValidationError):
        pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)


def test_concurrent_run_raises_lock_busy_error(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)

    with (
        statelock.acquire(config.state_dir),
        pytest.raises(LockBusyError),
    ):
        pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)


def test_lock_is_released_after_a_failed_run(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)

    with pytest.raises(ReportExportReferenceError):
        pipeline.run(
            config, run_date=date(2026, 7, 15), reference_path=tmp_path / "no-such-ref.csv.gz"
        )

    # The lock must have been released -- a fresh acquire must succeed.
    with statelock.acquire(config.state_dir):
        pass


def test_deliverable_replace_failure_raises_write_error(
    tmp_path: Path, reference_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # design.md §3.8 "交付檔重建保證": if the FINAL os.replace of the
    # deliverable fails (state already committed by then), the run must
    # still fail loudly with WriteError (exit 5) rather than silently
    # reporting success with a stale/missing deliverable. Only the
    # deliverable's own replace (targeting out_dir's *.xlsx) is made to
    # fail here -- state.commit()'s OWN internal os.replace (targeting
    # state_dir's records.csv, S9.1, which runs first) must still
    # succeed normally, so this exercises S9.2 specifically.
    real_replace = os.replace

    def _selective_raise(
        src: str | bytes | os.PathLike[str] | os.PathLike[bytes],
        dst: str | bytes | os.PathLike[str] | os.PathLike[bytes],
    ) -> None:
        if str(dst).endswith(".xlsx"):
            raise OSError(5, "Input/output error")
        real_replace(src, dst)

    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)

    monkeypatch.setattr(os, "replace", _selective_raise)
    with pytest.raises(WriteError, match="cannot finalize deliverable"):
        pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    # State-first ordering held: the batch WAS committed even though
    # the deliverable replace failed.
    result = state.load(config.state_dir)
    assert len(result.existing) == 1


# --------------------------------------------------------------------
# runs.jsonl audit trail gets one line per attempt (design.md §3.5.6)
# --------------------------------------------------------------------


def test_runs_jsonl_gets_one_entry_per_run(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1")])
    config = _config(tmp_path, input_path)

    pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)
    pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    lines = (config.state_dir / "runs.jsonl").read_text(encoding="utf-8").splitlines()
    assert len(lines) == 2
    first_record = json.loads(lines[0])
    assert first_record["appended"] == 1
    assert first_record["run_date"] == "2026-07-15"
    second_record = json.loads(lines[1])
    assert second_record["appended"] == 0
    assert second_record["skipped_cross"] == 1


def test_runs_jsonl_records_appended_request_ids(tmp_path: Path, reference_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    _write_input_csv(input_path, [_row("req-1"), _row("req-2")])
    config = _config(tmp_path, input_path)

    pipeline.run(config, run_date=date(2026, 7, 15), reference_path=reference_path)

    record = json.loads((config.state_dir / "runs.jsonl").read_text(encoding="utf-8"))
    assert record["appended_request_ids"] == ["req-1", "req-2"]


# --------------------------------------------------------------------
# Real anchor dataset -- template/source-log.csv (design.md §1.5.1, §1.5.2)
# --------------------------------------------------------------------

_TEMPLATE_INPUT = Path(__file__).resolve().parents[2] / "template" / "source-log.csv"
_REAL_REFERENCE = Path(__file__).resolve().parents[2] / "reference" / "hosp_id_map.csv.gz"


def test_real_template_input_produces_anchor_state_and_report(tmp_path: Path) -> None:
    config = Config(
        input_path=_TEMPLATE_INPUT, state_dir=tmp_path / "state", out_dir=tmp_path / "output"
    )

    summary = pipeline.run(config, run_date=date(2026, 7, 15), reference_path=_REAL_REFERENCE)

    assert summary.rows_in == 25
    assert summary.normal == 19
    assert summary.dropped_nonnormal == 6
    assert summary.new_records == 19
    assert summary.state_total == 19
    assert summary.unique_ips == 11
    assert summary.batch_seq == 1

    workbook = load_workbook(summary.deliverable)
    agg_sheet = workbook["院所分析"]
    counts = [row[3].value for row in agg_sheet.iter_rows(min_row=2, max_row=12)]
    assert counts == [1, 1, 1, 1, 1, 1, 3, 1, 7, 1, 1]
    assert sum(counts) == 19

    shuchuan = next(
        row for row in agg_sheet.iter_rows(min_row=2, max_row=12) if row[1].value == "0937010019"
    )
    assert shuchuan[2].value == "秀傳醫院"
    assert shuchuan[3].value == 7

    records_sheet = workbook["調閱紀錄"]
    assert records_sheet.max_row == 20  # header + 19 NORMAL rows
    # All 19 rows are batch 1 (the only batch) -> all highlighted.
    for row_idx in range(2, 21):
        assert records_sheet[f"A{row_idx}"].fill.patternType == "solid"
