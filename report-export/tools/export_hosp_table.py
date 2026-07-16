#!/usr/bin/env python3
"""Export HOSP_ID -> HOSP_ABBR from the source template into a
runtime-friendly `reference/hosp_id_map.csv.gz` + manifest.

One-off dev/ops tool (design.md §3.3): NOT part of the `report_export`
runtime package, NOT bundled into the runtime Docker image (design.md
§4.7.2 .dockerignore excludes `tools/`). Run once at build time and
again whenever the source template's master table changes (design.md
§3.3 update procedure).

Usage:
    python3 tools/export_hosp_table.py \
        [--source template/連線紀錄模板.xlsx] \
        [--out-dir reference]

Fail-loud (design.md §3.3): exits non-zero and prints to stderr if the
exported table does not match the shape measured against the real
template file and locked in design.md §1.5.8:

    row_count == 93781, dup_keys == 0, blank_abbr == 0,
    key_len_hist == {10: 93781}, leading_zero == 531.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import json
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Final

import openpyxl

TOOL_VERSION: Final[str] = "1.0.0"
SHEET_NAME: Final[str] = "HOSP_ID對照表"
EXPECTED_HEADER: Final[tuple[str, str]] = ("HOSP_ID", "HOSP_ABBR")

# Fail-loud expectations locked from design.md §1.5.8 (measured against
# the real committed template/連線紀錄模板.xlsx).
EXPECTED_ROW_COUNT: Final[int] = 93781
EXPECTED_DUP_KEYS: Final[int] = 0
EXPECTED_BLANK_ABBR: Final[int] = 0
EXPECTED_KEY_LEN_HIST: Final[dict[int, int]] = {10: 93781}
EXPECTED_LEADING_ZERO: Final[int] = 531

_REPORT_EXPORT_ROOT: Final[Path] = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE: Final[Path] = _REPORT_EXPORT_ROOT / "template" / "連線紀錄模板.xlsx"
DEFAULT_OUT_DIR: Final[Path] = _REPORT_EXPORT_ROOT / "reference"


class ExportValidationError(RuntimeError):
    """Raised when the source sheet or the exported table fails a fail-loud assertion."""


@dataclass(frozen=True, slots=True)
class ExportStats:
    row_count: int
    dup_keys: int
    blank_abbr: int
    key_len_hist: dict[int, int]
    leading_zero: int


def read_hosp_rows(source: Path) -> list[tuple[str, str]]:
    """Stream (HOSP_ID, HOSP_ABBR) pairs from the template's master sheet.

    Uses openpyxl `read_only` mode (design.md §3.3: a 2.3MB workbook
    with 93,781 data rows must never be loaded eagerly). Every cell is
    coerced to `str` on read so leading zeros in HOSP_ID survive
    regardless of the source cell's stored data type.
    """
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ExportValidationError(
                f"sheet {SHEET_NAME!r} not found in {source} (have: {workbook.sheetnames})"
            )
        sheet = workbook[SHEET_NAME]
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None or tuple(header[:2]) != EXPECTED_HEADER:
            raise ExportValidationError(
                f"unexpected header {header!r} in sheet {SHEET_NAME!r}; "
                f"expected {EXPECTED_HEADER!r}"
            )

        rows: list[tuple[str, str]] = []
        for line_no, row in enumerate(rows_iter, start=2):
            if row is None or all(cell is None for cell in row):
                continue
            hosp_id_raw, hosp_abbr_raw = row[0], row[1]
            if hosp_id_raw is None:
                raise ExportValidationError(f"row {line_no}: HOSP_ID is blank")
            # Deliberately NOT stripped: this is an exact-reproduction
            # export, not a cleanup pass. The master table's key_len_hist
            # is measured (design.md §1.5.8) at a uniform 10 chars for
            # EVERY key, including 6 sentinel/placeholder rows (e.g.
            # "TEST_USER ", "373104012 ") whose trailing space is part
            # of the stored 10-char key -- .strip()-ing it would corrupt
            # those keys to 9 chars and break the fail-loud row_count/
            # key_len_hist/leading_zero assertions below.
            hosp_id = str(hosp_id_raw)
            hosp_abbr = "" if hosp_abbr_raw is None else str(hosp_abbr_raw)
            rows.append((hosp_id, hosp_abbr))
        return rows
    finally:
        workbook.close()


def compute_stats(rows: list[tuple[str, str]]) -> ExportStats:
    key_counts = Counter(hosp_id for hosp_id, _ in rows)
    dup_keys = sum(1 for count in key_counts.values() if count > 1)
    blank_abbr = sum(1 for _, hosp_abbr in rows if not hosp_abbr)
    key_len_hist = dict(Counter(len(hosp_id) for hosp_id, _ in rows))
    leading_zero = sum(1 for hosp_id, _ in rows if hosp_id.startswith("0"))
    return ExportStats(
        row_count=len(rows),
        dup_keys=dup_keys,
        blank_abbr=blank_abbr,
        key_len_hist=key_len_hist,
        leading_zero=leading_zero,
    )


def validate_stats(stats: ExportStats) -> None:
    """Fail-loud assertions locked from design.md §1.5.8 / §3.3."""
    violations: list[str] = []
    if stats.row_count != EXPECTED_ROW_COUNT:
        violations.append(f"row_count={stats.row_count} != {EXPECTED_ROW_COUNT}")
    if stats.dup_keys != EXPECTED_DUP_KEYS:
        violations.append(f"dup_keys={stats.dup_keys} != {EXPECTED_DUP_KEYS}")
    if stats.blank_abbr != EXPECTED_BLANK_ABBR:
        violations.append(f"blank_abbr={stats.blank_abbr} != {EXPECTED_BLANK_ABBR}")
    if stats.key_len_hist != EXPECTED_KEY_LEN_HIST:
        violations.append(f"key_len_hist={stats.key_len_hist} != {EXPECTED_KEY_LEN_HIST}")
    if stats.leading_zero != EXPECTED_LEADING_ZERO:
        violations.append(f"leading_zero={stats.leading_zero} != {EXPECTED_LEADING_ZERO}")
    if violations:
        raise ExportValidationError(
            "exported hosp_id_map failed fail-loud validation: " + "; ".join(violations)
        )


def serialize_csv_bytes(rows: list[tuple[str, str]]) -> bytes:
    """Serialize (HOSP_ID, HOSP_ABBR) rows into the exact uncompressed
    CSV byte stream this tool writes to hosp_id_map.csv.gz (header +
    all data rows; `csv.QUOTE_MINIMAL`, csv module's default '\\r\\n'
    line terminator), independent of gzip framing.

    Single source of truth for "what are the uncompressed CSV bytes":
    both `write_csv_gz` (what actually lands on disk, gzip-wrapped)
    and `sha256_of_table` (what the manifest asserts) serialize
    through this one function, so the on-disk bytes and the manifest
    digest can never drift apart.
    """
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPECTED_HEADER)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def sha256_of_table(rows: list[tuple[str, str]]) -> str:
    """sha256 over the uncompressed CSV bytes (header + all data rows).

    Hashes the exact byte stream `write_csv_gz` writes into
    hosp_id_map.csv.gz (see `serialize_csv_bytes`) -- not a
    hand-rolled reconstruction -- so `gunzip -c hosp_id_map.csv.gz |
    sha256sum` reproduces this manifest field exactly. Independent of
    gzip framing (container/header/mtime), so it stays stable across
    gzip implementations/mtimes -- mirrors the "sha256 over
    header+all data rows" convention used for records.csv's own
    `#META` tail (design.md §3.5.3).
    """
    return hashlib.sha256(serialize_csv_bytes(rows)).hexdigest()


def write_csv_gz(rows: list[tuple[str, str]], dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_name(dest.name + ".tmp")
    payload = serialize_csv_bytes(rows)
    with gzip.open(tmp, mode="wb") as fh:
        fh.write(payload)
    tmp.replace(dest)


def write_manifest(*, source: Path, stats: ExportStats, sha256: str, dest: Path) -> None:
    sorted_key_len_hist = sorted(stats.key_len_hist.items())
    manifest = {
        "source": str(source),
        "exported_utc": datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "row_count": stats.row_count,
        "sha256": sha256,
        "key_len_hist": {str(length): count for length, count in sorted_key_len_hist},
        "dup_keys": stats.dup_keys,
        "blank_abbr": stats.blank_abbr,
        "leading_zero": stats.leading_zero,
        "tool_version": TOOL_VERSION,
    }
    dest.parent.mkdir(parents=True, exist_ok=True)
    tmp = dest.with_name(dest.name + ".tmp")
    payload = json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    tmp.write_text(payload, encoding="utf-8")
    tmp.replace(dest)


def run(source: Path, out_dir: Path) -> ExportStats:
    rows = read_hosp_rows(source)
    stats = compute_stats(rows)
    validate_stats(stats)
    sha256 = sha256_of_table(rows)
    write_csv_gz(rows, out_dir / "hosp_id_map.csv.gz")
    manifest_dest = out_dir / "hosp_id_map.manifest.json"
    write_manifest(source=source, stats=stats, sha256=sha256, dest=manifest_dest)
    return stats


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE,
        help=f"source template xlsx (default: {DEFAULT_SOURCE})",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=DEFAULT_OUT_DIR,
        help=f"output directory for hosp_id_map.csv.gz + manifest (default: {DEFAULT_OUT_DIR})",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        stats = run(args.source, args.out_dir)
    except (ExportValidationError, OSError) as exc:
        print(f"FATAL: {exc}", file=sys.stderr)
        return 1
    print(
        f"OK: exported {stats.row_count} rows -> {args.out_dir / 'hosp_id_map.csv.gz'} "
        f"(dup_keys={stats.dup_keys}, blank_abbr={stats.blank_abbr}, "
        f"leading_zero={stats.leading_zero}, key_len_hist={stats.key_len_hist})",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
